"""
Extract échéancier data from Excel and generate finance_echeancier.js

Reads: Echancier des projets.xlsx (35 sheets, one per project)
Generates: frontend/src/data/finance_echeancier.js
"""

import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

# --- Config ---
EXCEL_PATH = (
    r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau"
    r"\Fichiers de DOSSIER DASHBOARD - Data_Dashbords"
    r"\07_Finance\Echancier des projets.xlsx"
)
CREANCES_PATH = (
    r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau"
    r"\Fichiers de DOSSIER DASHBOARD - Data_Dashbords"
    r"\07_Finance\CREANCES Filatex SA+TCM.xlsx"
)
OUTPUT_JS = os.path.join(
    os.path.dirname(__file__),
    "frontend", "src", "data", "finance_echeancier.js",
)

SKIP_SHEETS = {"RECAP ENCAISSEMENTS", "TS", "5 VALH 1 VPAT"}
CURRENT_YEAR = 2026
MAX_COL = 250  # Safety cap for column scanning

MONTH_NAMES_FR = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "août": 8, "aout": 8,
    "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12, "decembre": 12,
    # Short forms
    "janv": 1, "jan": 1, "fév": 2, "fev": 2, "mar": 3, "avr": 4,
    "jui": 6, "juil": 7, "jul": 7, "sep": 9, "sept": 9,
    "oct": 10, "nov": 11, "déc": 12, "dec": 12,
}


def load_retard_clients():
    """Read CREANCES Excel — all clients listed there are 'en retard'.

    Sources:
    - 'TCM observ' sheet: column A = client names (rows 5+)
    - 'FLX observ' sheet: column B = client names (rows 3+)
    """
    retard_set = set()
    try:
        wb = openpyxl.load_workbook(CREANCES_PATH, data_only=True, read_only=True)

        # TCM observ — client names in column A, data starts row 5
        if "TCM observ" in wb.sheetnames:
            ws = wb["TCM observ"]
            for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
                name = row[0]
                if name and isinstance(name, str) and name.strip():
                    retard_set.add(name.strip().upper())

        # FLX observ — client names in column B, data starts row 3
        if "FLX observ" in wb.sheetnames:
            ws = wb["FLX observ"]
            for row in ws.iter_rows(min_row=3, min_col=2, max_col=2, values_only=True):
                name = row[0]
                if name and isinstance(name, str) and name.strip():
                    retard_set.add(name.strip().upper())

        wb.close()
        print(f"  Loaded {len(retard_set)} clients en retard from CREANCES Excel")

    except FileNotFoundError:
        print("  [WARN] CREANCES Excel not found, retard detection disabled")
    except Exception as e:
        print(f"  [WARN] Error reading CREANCES Excel: {e}")

    return retard_set


def normalize_text(text):
    """Normalize text: lowercase, strip, remove accents and special chars."""
    import unicodedata
    t = str(text).strip().lower()
    # Normalize unicode
    t = unicodedata.normalize("NFKD", t)
    # Remove combining diacritical marks
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t


def parse_month(text):
    """Try to parse a month name (French) to month number."""
    if not text:
        return None
    t = normalize_text(text)
    if not t or not t[0].isalpha():
        return None
    for name, num in MONTH_NAMES_FR.items():
        name_clean = normalize_text(name)
        if t.startswith(name_clean) or name_clean.startswith(t):
            return num
    return None


def is_total_row(row_vals):
    """Check if a row is a total/summary row."""
    for v in row_vals[:3]:
        if v and isinstance(v, str) and "TOTAL" in v.upper():
            return True
    return False


def safe_num(val):
    """Convert value to float, treating None and non-numeric as 0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(" ", "").replace(",", "."))
        except ValueError:
            return 0.0
    return 0.0


def period_key(year, month):
    """Create period key: 'YYYY' for non-current years, 'YYYY-MM' for current year."""
    if year == CURRENT_YEAR:
        return f"{year}-{month:02d}"
    return str(year)


def process_sheet(ws, retard_clients):
    """Process a single project sheet and return project data."""
    project_name = ws.title.strip()

    # Find header row with "CLIENT" in columns A-C
    header_row = None
    client_col = 2  # Default: column B
    for r in range(1, 11):
        for c in [1, 2, 3]:
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and "CLIENT" in val.upper():
                header_row = r
                client_col = c
                break
        if header_row:
            break

    if header_row is None:
        print(f"  [SKIP] {project_name}: no CLIENT header found in rows 1-10")
        return None

    year_row = header_row - 1

    # Find PRIX column (column with PRIX+AR/MGA — Ariary, not EUR)
    prix_col = None
    prix_candidates = []  # (priority, col)
    for c in range(1, min(ws.max_column + 1, MAX_COL)):
        val = ws.cell(row=header_row, column=c).value
        if val and isinstance(val, str):
            upper = val.upper()
            # Best: "PRIX Ar/TTC" or "Px tot MGA ttc"
            if ("PRIX" in upper or "PX" in upper) and ("AR" in upper or "MGA" in upper) and "EUR" not in upper and "USD" not in upper:
                prix_candidates.append((0, c))
            # Good: "Px tot Eur ttc" (will use EUR value if no Ariary col)
            elif ("PRIX" in upper or "PX" in upper) and ("EUR" in upper or "USD" in upper):
                prix_candidates.append((2, c))
            elif "PRIX" in upper and "TTC" in upper:
                prix_candidates.append((1, c))

    if prix_candidates:
        prix_candidates.sort()
        prix_col = prix_candidates[0][1]
    else:
        # Fallback
        for fallback in [9, 8, 7]:
            v = ws.cell(row=header_row, column=fallback).value
            if v:
                prix_col = fallback
                break
        if prix_col is None:
            prix_col = 9
        print(f"  [INFO] {project_name}: PRIX column not found, using col {prix_col}")

    # Find "Mens" column to know where monthly data starts
    mens_col = None
    for c in range(1, min(ws.max_column + 1, MAX_COL)):
        val = ws.cell(row=header_row, column=c).value
        if val and isinstance(val, str) and "MENS" in val.upper():
            mens_col = c
            break

    if mens_col is None:
        # Try to find first monthly column after prix_col
        for c in range(prix_col + 1, min(ws.max_column + 1, MAX_COL)):
            val = ws.cell(row=header_row, column=c).value
            if val and parse_month(val) is not None:
                mens_col = c - 1  # Monthly data starts after this
                break

    if mens_col is None:
        # Last resort: look for month names anywhere after client_col
        for c in range(client_col + 1, min(ws.max_column + 1, MAX_COL)):
            val = ws.cell(row=header_row, column=c).value
            if val and parse_month(val) is not None:
                mens_col = c - 1
                break

    if mens_col is None:
        print(f"  [SKIP] {project_name}: no monthly columns found")
        return None

    # Build column -> (year, month) mapping
    col_to_date = {}
    current_year_val = None

    # Scan from column 1 to capture year markers that appear before monthly cols
    for c in range(1, min(ws.max_column + 1, MAX_COL)):
        # Check year row for year marker
        year_val = ws.cell(row=year_row, column=c).value
        if year_val is not None:
            try:
                y = int(float(str(year_val)))
                if 2000 <= y <= 2050:
                    current_year_val = y
            except (ValueError, TypeError):
                pass

        # Only map columns after mens_col
        if c > mens_col:
            month_val = ws.cell(row=header_row, column=c).value
            month_num = parse_month(month_val)
            if month_num is not None and current_year_val is not None:
                col_to_date[c] = (current_year_val, month_num)

    if not col_to_date:
        print(f"  [SKIP] {project_name}: no date columns mapped")
        return None

    # Process client rows
    clients = []
    for r in range(header_row + 1, ws.max_row + 1):
        # Check first few cols for TOTAL marker
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max(prix_col + 2, 6))]

        # Skip total rows (check all visible cols for "TOTAL", "ENCAISSEMENTS", "RESTE")
        skip = False
        for v in row_vals:
            if v and isinstance(v, str):
                vu = v.upper().strip()
                if vu.startswith("TOTAL") or vu.startswith("ENCAISSEMENTS") or vu.startswith("RESTE"):
                    skip = True
                    break
        if skip:
            continue

        client_name = ws.cell(row=r, column=client_col).value
        if client_name is None or (isinstance(client_name, str) and client_name.strip() == ""):
            continue

        client_name = str(client_name).strip()

        # Skip if client name looks like a header, section label, or metadata
        cu = client_name.upper()
        if cu in ("CLIENT", "NOM", ""):
            continue
        if any(kw in cu for kw in [
            "MONTANT TOTAL", "COURS", "1EUR", "1 EUR",
            "ENCAISSEMENT", "RESTE A", "PRIX TOTAL",
        ]):
            continue

        montant = safe_num(ws.cell(row=r, column=prix_col).value)

        # Check retard status
        en_retard = client_name.upper() in retard_clients

        # Collect monthly payments
        payments = defaultdict(float)
        for c, (year, month) in col_to_date.items():
            val = safe_num(ws.cell(row=r, column=c).value)
            if val != 0:
                pk = period_key(year, month)
                payments[pk] += val

        clients.append({
            "client": client_name,
            "montantContractuel": montant,
            "enRetard": en_retard,
            "payments": dict(payments),
        })

    if not clients:
        print(f"  [SKIP] {project_name}: no client rows found")
        return None

    # Sort clients by montantContractuel descending
    clients.sort(key=lambda x: x["montantContractuel"], reverse=True)

    # Build project timeline (aggregate all clients)
    timeline_data = defaultdict(lambda: {"contractuel": 0.0, "aTemps": 0.0, "enRetard": 0.0})
    total_contractuel = 0.0
    nb_en_retard = 0
    nb_a_temps = 0

    for cl in clients:
        total_contractuel += cl["montantContractuel"]
        if cl["enRetard"]:
            nb_en_retard += 1
        else:
            nb_a_temps += 1

        for pk, amount in cl["payments"].items():
            timeline_data[pk]["contractuel"] += amount
            if cl["enRetard"]:
                timeline_data[pk]["enRetard"] += amount
            else:
                timeline_data[pk]["aTemps"] += amount

    # Sort timeline periods
    def sort_period(p):
        if "-" in p:
            parts = p.split("-")
            return (int(parts[0]), int(parts[1]))
        return (int(p), 0)

    sorted_periods = sorted(timeline_data.keys(), key=sort_period)
    timeline = []
    for pk in sorted_periods:
        d = timeline_data[pk]
        if d["contractuel"] == 0 and d["aTemps"] == 0 and d["enRetard"] == 0:
            continue
        timeline.append({
            "periode": pk,
            "contractuel": round(d["contractuel"], 2),
            "aTemps": round(d["aTemps"], 2),
            "enRetard": round(d["enRetard"], 2),
        })

    # Build client list (without payments dict for output)
    client_output = []
    for cl in clients:
        client_output.append({
            "client": cl["client"],
            "montantContractuel": cl["montantContractuel"],
            "enRetard": cl["enRetard"],
        })

    return {
        "projet": project_name,
        "totalContractuel": round(total_contractuel, 2),
        "nbClients": len(clients),
        "nbEnRetard": nb_en_retard,
        "nbATemps": nb_a_temps,
        "clients": client_output,
        "timeline": timeline,
        "_payments_by_client": clients,  # Keep for global aggregation
    }


def build_global_timeline(projects):
    """Aggregate timeline across all projects."""
    global_data = defaultdict(lambda: {"contractuel": 0.0, "aTemps": 0.0, "enRetard": 0.0})

    for proj in projects:
        for cl in proj["_payments_by_client"]:
            for pk, amount in cl["payments"].items():
                global_data[pk]["contractuel"] += amount
                if cl["enRetard"]:
                    global_data[pk]["enRetard"] += amount
                else:
                    global_data[pk]["aTemps"] += amount

    def sort_period(p):
        if "-" in p:
            parts = p.split("-")
            return (int(parts[0]), int(parts[1]))
        return (int(p), 0)

    sorted_periods = sorted(global_data.keys(), key=sort_period)
    timeline = []
    for pk in sorted_periods:
        d = global_data[pk]
        if d["contractuel"] == 0 and d["aTemps"] == 0 and d["enRetard"] == 0:
            continue
        timeline.append({
            "periode": pk,
            "contractuel": round(d["contractuel"], 2),
            "aTemps": round(d["aTemps"], 2),
            "enRetard": round(d["enRetard"], 2),
        })
    return timeline


def main():
    print(f"Loading retard clients from finance_data.js...")
    retard_clients = load_retard_clients()
    print(f"  Found {len(retard_clients)} clients with retardJours > 0")

    print(f"\nLoading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"  Found {len(wb.sheetnames)} sheets")

    projects = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in SKIP_SHEETS:
            print(f"  [SKIP] {sheet_name}: in skip list")
            continue

        ws = wb[sheet_name]
        result = process_sheet(ws, retard_clients)
        if result:
            projects.append(result)
            print(
                f"  [OK] {result['projet']}: "
                f"{result['nbClients']} clients, "
                f"{result['nbEnRetard']} en retard, "
                f"total={result['totalContractuel']:,.0f}"
            )

    wb.close()

    # Sort projects alphabetically
    projects.sort(key=lambda x: x["projet"].upper())

    # Build global timeline
    global_timeline = build_global_timeline(projects)

    # Clean up internal data before output
    for proj in projects:
        del proj["_payments_by_client"]

    # Write JS output
    print(f"\nWriting {OUTPUT_JS}...")
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write("// Auto-generated from Echancier des projets.xlsx\n")
        f.write("export const TCM_ECHEANCIER = ")
        f.write(json.dumps(projects, indent=2, ensure_ascii=False))
        f.write(";\n\n")
        f.write("export const TCM_ECHEANCIER_GLOBAL = ")
        f.write(json.dumps(global_timeline, indent=2, ensure_ascii=False))
        f.write(";\n")

    # Summary
    total_clients = sum(p["nbClients"] for p in projects)
    total_retard = sum(p["nbEnRetard"] for p in projects)
    total_contractuel = sum(p["totalContractuel"] for p in projects)

    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"  Projects:         {len(projects)}")
    print(f"  Total clients:    {total_clients}")
    print(f"  En retard:        {total_retard}")
    print(f"  A temps:          {total_clients - total_retard}")
    print(f"  Total contractuel: {total_contractuel:,.0f} Ar")
    print(f"  Timeline periods:  {len(global_timeline)}")
    print(f"  Output:           {OUTPUT_JS}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
