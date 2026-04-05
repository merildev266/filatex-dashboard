import os
import openpyxl
from datetime import datetime

CAPEX_PATH = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - GROUPE FILATEX",
    "Bureau",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "04_CAPEX",
    "CAPEX.xlsx",
)
wb = openpyxl.load_workbook(CAPEX_PATH, data_only=True)

def sv(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return round(v, 2)
    if isinstance(v, str) and v.strip() in ['', '-', '#REF!']: return 0
    try: return round(float(v), 2)
    except (ValueError, TypeError): return 0

def fmtAmt(n):
    if n == 0: return '\u2014'
    if abs(n) >= 1000000:
        v = n / 1000000
        if v == int(v): return f'{int(v)} M$'
        return f'{v:.1f} M$'
    if abs(n) >= 1000:
        return f'{int(round(n/1000))} k$'
    return f'{int(n)} $'

def fmtAmtFull(n):
    if n == 0: return '\u2014'
    s = f'{int(round(n)):,}'.replace(',', ' ')
    return f'{s} $'

def fmtDate(v):
    if v is None: return '\u2014'
    if hasattr(v, 'strftime'): return v.strftime('%d.%m.%y')
    if isinstance(v, str):
        if '/' in v: return v
        if v in ['N/A', '-', '']: return '\u2014'
    return str(v)[:10] if v else '\u2014'

def fmtPct(n):
    if n == 0: return '\u2014'
    if n < 1: return f'{n*100:.0f}%'
    return f'{n:.0f}%'

# === ENAT ===
ws = wb['ENAT ']
enat_projects = []
for r in range(9, 37):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL': continue
    init = sv(ws.cell(r,3).value)
    incurred = sv(ws.cell(r,5).value)
    h1_26 = sv(ws.cell(r,6).value)
    h2_26 = sv(ws.cell(r,7).value)
    y2027 = sv(ws.cell(r,8).value)
    tri = sv(ws.cell(r,10).value)
    start = ws.cell(r,12).value
    end_init = ws.cell(r,13).value
    end_rev = ws.cell(r,14).value
    enat_projects.append({
        'name': name, 'init': init, 'incurred': incurred,
        'h1_26': h1_26, 'h2_26': h2_26, 'y2027': y2027,
        'tri': tri, 'start': start, 'end_init': end_init, 'end_rev': end_rev
    })

# Group studies with main projects
groups = {}
order = []
for p in enat_projects:
    n = p['name']
    is_study = '(Studies)' in n or '(Study)' in n
    base = n.replace(' (Studies)', '').replace('(Studies)', '').replace(' (Study)', '').replace('(Study)', '').strip()
    if base not in groups:
        groups[base] = {'main': None, 'study': None}
        order.append(base)
    if is_study:
        groups[base]['study'] = p
    else:
        groups[base]['main'] = p

enr_projects = []
idx = 1
for base in order:
    g = groups[base]
    main = g['main'] or g['study']
    study = g['study']
    if not main: continue
    total_init = main['init'] + (study['init'] if study else 0)
    total_incurred = main['incurred'] + (study['incurred'] if study else 0)
    if total_init == 0 and total_incurred == 0: continue

    budget = total_init  # ENAT has no revisions
    pct = round(total_incurred / budget * 100) if budget > 0 else 0
    pct = min(pct, 100)

    status = 'on-track'
    end_rev = main.get('end_rev')
    if end_rev and isinstance(end_rev, str) and len(str(end_rev)) > 15:
        status = 'delayed'

    tri_val = main['tri']
    h1 = main['h1_26'] + (study['h1_26'] if study else 0)
    h2 = main['h2_26'] + (study['h2_26'] if study else 0)

    # Generate simple cashflow indicators
    total_cf = h1 + h2 + main['y2027']
    if total_cf > 0:
        m1 = round(h1/6/total_cf*60) if h1 else 0
        m2 = round(h2/6/total_cf*60) if h2 else 0
        cf_out = [max(1,m1)]*3 + [max(1,m2)]*3 if (m1+m2) > 0 else [0]*6
    else:
        cf_out = [0]*6
    cf_in = [0]*6

    enr_projects.append({
        'id': f'enr-{idx}', 'name': base, 'status': status,
        'investInit': fmtAmtFull(total_init),
        'investReel': fmtAmtFull(total_incurred) if total_incurred > 0 else '\u2014',
        'deltaInvest': '\u2014',
        'etatEnCours': fmtAmt(total_incurred), 'etatTotal': fmtAmt(budget), 'etatPct': pct,
        'cfIn': cf_in, 'cfOut': cf_out, 'cfNet': '\u2014',
        'triInit': fmtPct(tri_val) if tri_val else '\u2014', 'triReel': '\u2014', 'deltaPerf': 'up',
        'dateDebInit': fmtDate(main['start']), 'dateDebReel': fmtDate(main['start']),
        'dateFinInit': fmtDate(main['end_init']),
        'dateFinReel': fmtDate(end_rev) if (end_rev and hasattr(end_rev, 'strftime')) else '\u2014',
    })
    idx += 1

# Add LIDERA projects
ws = wb['LIDERA ']
for r in range(9, 18):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL': continue
    init = sv(ws.cell(r,3).value)
    revised = sv(ws.cell(r,4).value)
    incurred = sv(ws.cell(r,5).value)
    h1_26 = sv(ws.cell(r,6).value)
    h2_26 = sv(ws.cell(r,7).value)
    y2027 = sv(ws.cell(r,8).value)
    tri = sv(ws.cell(r,10).value)
    start = ws.cell(r,12).value
    end_init = ws.cell(r,13).value
    end_rev = ws.cell(r,14).value
    budget = revised if revised > 0 else init
    if budget == 0 and incurred == 0: continue
    pct = round(incurred / budget * 100) if budget > 0 else 0
    pct = min(pct, 100)
    delta = f'{round((revised-init)/init*100):+d}%' if init > 0 and revised > 0 else '\u2014'

    enr_projects.append({
        'id': f'enr-{idx}',
        'name': 'LIDERA \u2013 ' + name.strip(),
        'status': 'on-track',
        'investInit': fmtAmtFull(init) if init > 0 else '\u2014',
        'investReel': fmtAmtFull(revised) if revised > 0 else '\u2014',
        'deltaInvest': delta,
        'etatEnCours': fmtAmt(incurred), 'etatTotal': fmtAmt(budget), 'etatPct': pct,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': fmtPct(tri) if tri else '\u2014', 'triReel': '\u2014', 'deltaPerf': 'up',
        'dateDebInit': fmtDate(start), 'dateDebReel': fmtDate(start),
        'dateFinInit': fmtDate(end_init), 'dateFinReel': fmtDate(end_rev) if end_rev else '\u2014',
    })
    idx += 1

# Add LIDERA SERV
ws = wb['LIDERA SERV ']
for r in range(9, 12):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL': continue
    init = sv(ws.cell(r,3).value)
    revised = sv(ws.cell(r,4).value)
    incurred = sv(ws.cell(r,5).value)
    tri_init = sv(ws.cell(r,10).value)
    tri_rev = sv(ws.cell(r,11).value)
    budget = revised if revised > 0 else init
    if budget == 0: continue
    pct = round(incurred / budget * 100) if budget > 0 else 0
    short = name[:50]
    enr_projects.append({
        'id': f'enr-{idx}',
        'name': 'LIDERA SERV \u2013 ' + short.strip(),
        'status': 'on-track',
        'investInit': fmtAmtFull(init),
        'investReel': fmtAmtFull(revised) if revised != init else '\u2014',
        'deltaInvest': '\u2014',
        'etatEnCours': fmtAmt(incurred), 'etatTotal': fmtAmt(budget), 'etatPct': pct,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': fmtPct(tri_init) if tri_init else '\u2014',
        'triReel': fmtPct(tri_rev) if tri_rev else '\u2014',
        'deltaPerf': 'up' if tri_rev >= tri_init else 'down' if tri_rev > 0 else 'up',
        'dateDebInit': '\u2014', 'dateDebReel': '\u2014', 'dateFinInit': '\u2014', 'dateFinReel': '\u2014',
    })
    idx += 1

# === HFO ===
ws = wb['HFO ']
hfo_projects = []
idx = 1
current_section = None
section_data = {}
section_order = []

for r in range(9, 57):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL' or name.strip() == '': continue

    c4 = sv(ws.cell(r,4).value)
    c5 = sv(ws.cell(r,5).value)
    c6 = sv(ws.cell(r,6).value)
    c7 = sv(ws.cell(r,7).value)
    c8 = sv(ws.cell(r,8).value)
    c13 = ws.cell(r,13).value
    c14 = ws.cell(r,14).value

    name_str = str(name).strip()

    # Section headers contain "USD" or "IRR"
    if 'USD' in name_str or 'IRR' in name_str:
        clean = name_str.split(':')[0].strip() if ':' in name_str else name_str
        # Remove trailing markers
        for suffix in [' -', ' \u2013']:
            if clean.endswith(suffix): clean = clean[:-len(suffix)]
        current_section = clean

        total_named = 0
        irr_named = 0
        if 'USD' in name_str:
            try:
                p = name_str.split('USD')[1].split('/')[0].strip()
                p = p.replace(',', '').replace(' ', '').replace('\xa0', '')
                total_named = float(p)
            except (ValueError, IndexError): pass
        if 'IRR' in name_str:
            try:
                p = name_str.split('IRR')[1].strip().rstrip('%').strip()
                irr_named = float(p)
            except (ValueError, IndexError): pass

        section_data[current_section] = {
            'total_named': total_named, 'irr': irr_named,
            'total_init': 0, 'total_revised': 0, 'total_incurred': 0,
            'sub_items': [], 'start': None, 'end': None
        }
        section_order.append(current_section)
    elif current_section and current_section in section_data:
        sd = section_data[current_section]
        sd['total_init'] += c4
        sd['total_revised'] += c5
        sd['total_incurred'] += c6
        sd['sub_items'].append(name_str)
        if c13 and not sd['start']: sd['start'] = c13
        if c14 and not sd['end']: sd['end'] = c14

for sec_name in section_order:
    sd = section_data[sec_name]
    init_val = sd['total_named'] if sd['total_named'] > 0 else sd['total_init']
    revised_val = sd['total_revised']
    budget = revised_val if revised_val > 0 else init_val
    if budget == 0: continue

    incurred = sd['total_incurred']
    pct = round(incurred / budget * 100) if budget > 0 else 0
    pct = min(pct, 100)

    over = incurred > budget * 1.1

    hfo_projects.append({
        'id': f'hfo-{idx}', 'name': sec_name,
        'status': 'over-budget' if over else 'on-track',
        'investInit': fmtAmtFull(init_val) if init_val > 0 else '\u2014',
        'investReel': fmtAmtFull(revised_val) if revised_val > 0 and revised_val != init_val else '\u2014',
        'deltaInvest': '\u2014',
        'etatEnCours': fmtAmt(incurred), 'etatTotal': fmtAmt(budget), 'etatPct': pct,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': f"{int(sd['irr'])}%" if sd['irr'] else '\u2014', 'triReel': '\u2014', 'deltaPerf': 'up',
        'dateDebInit': fmtDate(sd['start']), 'dateDebReel': fmtDate(sd['start']),
        'dateFinInit': fmtDate(sd['end']), 'dateFinReel': '\u2014',
    })
    idx += 1

# === IMMO ===
immo_projects = []
idx = 1

ws = wb['IMMO TRAV']
for r in range(9, 33):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL' or name.strip() == '': continue
    init = sv(ws.cell(r,3).value)
    revised = sv(ws.cell(r,4).value)
    incurred = sv(ws.cell(r,5).value)
    tri = sv(ws.cell(r,10).value)
    start = ws.cell(r,12).value
    end_rev = ws.cell(r,14).value
    budget = revised if revised > 0 else init
    if budget == 0 and incurred == 0: continue
    pct = round(incurred / budget * 100) if budget > 0 else 0
    pct = min(pct, 100)
    delta = f'{round((revised-init)/init*100):+d}%' if init > 0 and revised > 0 and abs(revised - init) > 100 else '\u2014'

    immo_projects.append({
        'id': f'immo-{idx}', 'name': name.strip(),
        'status': 'on-track',
        'investInit': fmtAmtFull(init) if init > 0 else '\u2014',
        'investReel': fmtAmtFull(revised) if revised > 0 and abs(revised-init) > 100 else '\u2014',
        'deltaInvest': delta,
        'etatEnCours': fmtAmt(incurred), 'etatTotal': fmtAmt(budget), 'etatPct': pct,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': fmtPct(tri) if tri else '\u2014', 'triReel': '\u2014', 'deltaPerf': 'up',
        'dateDebInit': fmtDate(start), 'dateDebReel': fmtDate(start),
        'dateFinInit': '\u2014', 'dateFinReel': fmtDate(end_rev) if end_rev else '\u2014',
    })
    idx += 1

ws = wb['IMMO Foncier']
for r in range(9, 22):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL': continue
    total = sv(ws.cell(r,6).value)
    if total == 0: continue
    immo_projects.append({
        'id': f'immo-{idx}', 'name': 'Foncier \u2013 ' + name.strip(),
        'status': 'on-track',
        'investInit': fmtAmtFull(total), 'investReel': '\u2014', 'deltaInvest': '\u2014',
        'etatEnCours': '\u2014', 'etatTotal': fmtAmt(total), 'etatPct': 0,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': '\u2014', 'triReel': '\u2014', 'deltaPerf': 'up',
        'dateDebInit': '\u2014', 'dateDebReel': '\u2014', 'dateFinInit': '\u2014', 'dateFinReel': '\u2014',
    })
    idx += 1

# === VENTURES ===
ven_projects = []
idx = 1
ws = wb['VENTURES EXT']

# Read disbursement data from section 2 (R25-R34)
disbursements = {}
for r in range(25, 34):
    n = ws.cell(r,2).value
    if not n or n == 'TOTAL': continue
    disbursements[n.strip().lower()] = sv(ws.cell(r,5).value)

for r in range(9, 18):
    name = ws.cell(r,2).value
    if not name or name == 'TOTAL': continue
    init = sv(ws.cell(r,3).value)
    revised = sv(ws.cell(r,4).value)
    incurred = sv(ws.cell(r,5).value)
    tri_init = sv(ws.cell(r,10).value)
    tri_rev = sv(ws.cell(r,11).value)
    budget = revised if revised > 0 else init
    if budget == 0: continue

    # Find disbursement
    key = name.strip().rstrip(' :').lower()
    disbursed = 0
    for dk, dv in disbursements.items():
        if key[:8] in dk or dk[:8] in key:
            disbursed = dv
            break

    actual_amt = disbursed if disbursed > 0 else incurred
    actual_pct = round(actual_amt / budget * 100) if budget > 0 else 0
    actual_pct = min(actual_pct, 100)
    delta = f'{round((revised-init)/init*100):+d}%' if init > 0 and revised > 0 and abs(revised-init) > 100 else '\u2014'

    ven_projects.append({
        'id': f'ven-{idx}', 'name': name.strip().rstrip(' :'),
        'status': 'delayed' if 'restructure' in name.lower() else 'on-track',
        'investInit': fmtAmtFull(init),
        'investReel': fmtAmtFull(revised) if revised > 0 and abs(revised-init) > 100 else '\u2014',
        'deltaInvest': delta,
        'etatEnCours': fmtAmt(actual_amt), 'etatTotal': fmtAmt(budget), 'etatPct': actual_pct,
        'cfIn': [0]*6, 'cfOut': [0]*6, 'cfNet': '\u2014',
        'triInit': fmtPct(tri_init) if tri_init else '\u2014',
        'triReel': fmtPct(tri_rev) if tri_rev else '\u2014',
        'deltaPerf': 'up' if (tri_rev and tri_rev >= tri_init) else ('down' if tri_rev > 0 else 'up'),
        'dateDebInit': '\u2014', 'dateDebReel': '\u2014', 'dateFinInit': '\u2014', 'dateFinReel': '\u2014',
    })
    idx += 1

# === TOTALS ===
enr_init = 121953273 + 65200000 + 1288528
enr_revised = 121953273 + 30240386 + 826350
enr_incurred = 6357278.09 + 312000 + 0

hfo_init = 8599477
hfo_revised = 8309579
hfo_incurred = 2863362

immo_init = 155687965 + 625556 + 1840000
immo_revised = 102886975 + 628086 + 1840000
immo_incurred = 6078285 + 340478

ven_init = 30900000
ven_revised = 29722826
ven_disbursed = 19950000

total_revised = enr_revised + hfo_revised + immo_revised + ven_revised
total_init = enr_init + hfo_init + immo_init + ven_init
total_incurred = enr_incurred + hfo_incurred + immo_incurred + ven_disbursed

# === GENERATE JS ===
def esc(s):
    return s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")

def proj_js(p):
    lines = []
    lines.append(f"      {{")
    lines.append(f"        id:'{p['id']}', name:'{esc(p['name'])}', status:'{p['status']}',")
    lines.append(f"        investInit: '{esc(p['investInit'])}', investReel: '{esc(p['investReel'])}', deltaInvest: '{esc(p['deltaInvest'])}',")
    lines.append(f"        etatEnCours: '{esc(p['etatEnCours'])}', etatTotal: '{esc(p['etatTotal'])}', etatPct: {p['etatPct']},")
    lines.append(f"        cfIn: {p['cfIn']}, cfOut: {p['cfOut']}, cfNet: '{esc(p['cfNet'])}',")
    lines.append(f"        triInit: '{esc(p['triInit'])}', triReel: '{esc(p['triReel'])}', deltaPerf: '{p['deltaPerf']}',")
    lines.append(f"        dateDebInit: '{esc(p['dateDebInit'])}', dateDebReel: '{esc(p['dateDebReel'])}', dateFinInit: '{esc(p['dateFinInit'])}', dateFinReel: '{esc(p['dateFinReel'])}',")
    lines.append(f"      }}")
    return '\n'.join(lines)

js_parts = []
js_parts.append('const capexData = {')

# EnR
js_parts.append("  enr: {")
js_parts.append("    color: '#00ab63', colorRgb: '116,184,89',")
js_parts.append("    title: 'EnR \u2014 \u00c9nergies Renouvelables (ENAT + LIDERA)',")
js_parts.append("    projects: [")
js_parts.append(',\n'.join(proj_js(p) for p in enr_projects))
js_parts.append("    ]")
js_parts.append("  },")

# HFO
js_parts.append("  hfo: {")
js_parts.append("    color: '#426ab3', colorRgb: '100,136,255',")
js_parts.append("    title: 'HFO \u2014 Centrales Thermiques',")
js_parts.append("    projects: [")
js_parts.append(',\n'.join(proj_js(p) for p in hfo_projects))
js_parts.append("    ]")
js_parts.append("  },")

# IMMO
js_parts.append("  immo: {")
js_parts.append("    color: '#FDB823', colorRgb: '248,193,0',")
js_parts.append("    title: 'Immobilier \u2014 Patrimoine (Travaux + Foncier)',")
js_parts.append("    projects: [")
js_parts.append(',\n'.join(proj_js(p) for p in immo_projects))
js_parts.append("    ]")
js_parts.append("  },")

# Ventures
js_parts.append("  ventures: {")
js_parts.append("    color: '#f37056', colorRgb: '255,135,88',")
js_parts.append("    title: 'Ventures \u2014 Investissements Externes',")
js_parts.append("    projects: [")
js_parts.append(',\n'.join(proj_js(p) for p in ven_projects))
js_parts.append("    ]")
js_parts.append("  }")
js_parts.append("};")

js = '\n'.join(js_parts)

with open('capex_generated.js', 'w', encoding='utf-8') as f:
    f.write(js)

# Print summary
print("=== TOTALS ===")
print(f"EnR: init={fmtAmt(enr_init)}, revised={fmtAmt(enr_revised)}, incurred={fmtAmt(enr_incurred)}, pct={enr_incurred/enr_revised*100:.1f}%")
print(f"HFO: init={fmtAmt(hfo_init)}, revised={fmtAmt(hfo_revised)}, incurred={fmtAmt(hfo_incurred)}, pct={hfo_incurred/hfo_revised*100:.1f}%")
print(f"IMMO: init={fmtAmt(immo_init)}, revised={fmtAmt(immo_revised)}, incurred={fmtAmt(immo_incurred)}, pct={immo_incurred/immo_revised*100:.1f}%")
print(f"VEN: init={fmtAmt(ven_init)}, revised={fmtAmt(ven_revised)}, disbursed={fmtAmt(ven_disbursed)}, pct={ven_disbursed/ven_revised*100:.1f}%")
print(f"\nGRAND TOTAL: revised={fmtAmt(total_revised)}, init={fmtAmt(total_init)}, incurred={fmtAmt(total_incurred)}, pct={total_incurred/total_revised*100:.1f}%")
print(f"Delta init vs revised: {(total_revised-total_init)/total_init*100:+.1f}%")
print(f"\nProject counts: EnR={len(enr_projects)}, HFO={len(hfo_projects)}, IMMO={len(immo_projects)}, VEN={len(ven_projects)}")
print(f"Total projects: {len(enr_projects)+len(hfo_projects)+len(immo_projects)+len(ven_projects)}")
print(f"\nJS file size: {len(js)} chars")

# Section card numbers
print(f"\n=== SECTION CARD DATA ===")
print(f"EnR: {len(enr_projects)} projets, budget {fmtAmt(enr_revised)}, init {fmtAmt(enr_init)}, realise {fmtAmt(enr_incurred)}, {enr_incurred/enr_revised*100:.0f}%")
print(f"HFO: {len(hfo_projects)} projets, budget {fmtAmt(hfo_revised)}, init {fmtAmt(hfo_init)}, realise {fmtAmt(hfo_incurred)}, {hfo_incurred/hfo_revised*100:.0f}%")
print(f"IMMO: {len(immo_projects)} projets, budget {fmtAmt(immo_revised)}, init {fmtAmt(immo_init)}, realise {fmtAmt(immo_incurred)}, {immo_incurred/immo_revised*100:.0f}%")
print(f"VEN: {len(ven_projects)} projets, budget {fmtAmt(ven_revised)}, init {fmtAmt(ven_init)}, realise {fmtAmt(ven_disbursed)}, {ven_disbursed/ven_revised*100:.0f}%")
