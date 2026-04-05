"""Create Weekly Investments Avancement template (52 sheets S01-S52)."""
import openpyxl, sys, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\03_ Investments\Reporting\Weekly_Investments_Avancement_v2.xlsx"

PROJECTS = [
    ("oasis",            "OASIS",                  "Imtiaz", "Externe"),
    ("orga-earth",       "Orga Earth",             "Imtiaz", "Externe"),
    ("hakanto-house",    "Hakanto House",          "Imtiaz", "Externe"),
    ("mlf",              "MLF",                    "Imtiaz", "Externe"),
    ("energiestro",      "Energiestro",            "Imtiaz", "Externe"),
    ("hotel-tamatave",   "Hotel Tamatave",         "Imtiaz", "Externe"),
    ("sunfarming",       "SunFarming",             "Imtiaz", "Externe"),
    ("afridoctor",       "Afridoctor",             "Imtiaz", "Externe"),
    ("artemis",          "Artemis",                "Imtiaz", "Externe"),
    ("bgfi",             "BGFI",                   "Imtiaz", "Externe"),
    ("oui-coding",       "Oui Coding",             "Imtiaz", "Externe"),
    ("seed-star",        "Seed Star",              "Imtiaz", "Externe"),
    ("cafe-mary",        "Cafe Mary",              "Imtiaz", "Interne"),
    ("ghu",              "GHU",                    "Imtiaz", "Interne"),
    ("haya",             "Haya",                   "Imtiaz", "Interne"),
    ("maison-cotonniers","Maison des Cotonniers",  "Imtiaz", "Interne"),
    ("show-room",        "Show Room",              "Imtiaz", "Interne"),
    ("ssls",             "SSLS",                   "Imtiaz", "Interne"),
    ("taxi-brousse-pizza","Taxi Brousse Pizza",    "Imtiaz", "Interne"),
]

# Styles
DARK_BLUE  = PatternFill('solid', fgColor='1B2A4A')
MED_BLUE   = PatternFill('solid', fgColor='2C3E6B')
GRAY_HDR   = PatternFill('solid', fgColor='D9E1F2')
WHITE      = PatternFill('solid', fgColor='FFFFFF')
LIGHT_ALT  = PatternFill('solid', fgColor='F2F5FA')
SECTION_EXT = PatternFill('solid', fgColor='E8F0FE')
SECTION_INT = PatternFill('solid', fgColor='FFF3E0')

FONT_TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
FONT_INFO  = Font(name='Arial', size=10, color='FFFFFF')
FONT_HDR   = Font(name='Arial', size=9, bold=True, color='1B2A4A')
FONT_DATA  = Font(name='Arial', size=9, color='333333')
FONT_ID    = Font(name='Arial', size=9, color='666666')

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

HEADERS = [
    ("A", "ID Projet",          12),
    ("B", "Projet",             22),
    ("C", "Type",               10),
    ("D", "Statut",             14),
    ("E", "Avancement / Update",45),
    ("F", "Points de Blocage",  35),
    ("G", "Actions Prevus",     35),
    ("H", "Actions Realises",   40),
    ("I", "Mise a jour",        14),
    ("J", "Commentaires DG",    30),
    ("K", "Reponse",             30),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Week 1 Monday for 2026 (ISO): 2025-12-29
week1_mon = date(2025, 12, 29)

for wk_num in range(1, 53):
    sname = f"S{wk_num:02d}"
    ws = wb.create_sheet(sname)

    monday = week1_mon + timedelta(weeks=wk_num - 1)
    friday = monday + timedelta(days=4)

    # Title row
    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = "FILATEX VENTURES - Suivi Hebdomadaire Investments"
    c.font = FONT_TITLE
    c.fill = DARK_BLUE
    c.alignment = ALIGN_C
    for col in range(2, 12):
        ws.cell(1, col).fill = DARK_BLUE

    # Info row
    ws.merge_cells('A2:K2')
    ws['A2'].value = f"Semaine {wk_num}  |  {monday.strftime('%d/%m/%Y')} - {friday.strftime('%d/%m/%Y')}"
    ws['A2'].font = FONT_INFO
    ws['A2'].fill = MED_BLUE
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    for col in range(2, 12):
        ws.cell(2, col).fill = MED_BLUE

    # Row 3 spacer
    ws.row_dimensions[3].height = 6

    # --- SECTION: EXTERNE ---
    ws.merge_cells('A4:K4')
    ws['A4'].value = "INVESTISSEMENTS EXTERNES"
    ws['A4'].font = Font(name='Arial', size=10, bold=True, color='1565C0')
    ws['A4'].fill = SECTION_EXT
    ws['A4'].alignment = ALIGN_L
    for col in range(2, 12):
        ws.cell(4, col).fill = SECTION_EXT

    # Headers row 5
    for i, (col_letter, hdr_text, width) in enumerate(HEADERS):
        cell = ws.cell(5, i + 1)
        cell.value = hdr_text
        cell.font = FONT_HDR
        cell.fill = GRAY_HDR
        cell.alignment = ALIGN_C
        cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = width

    # Externe projects (rows 6+)
    row = 6
    ext_projects = [p for p in PROJECTS if p[3] == "Externe"]
    for pidx, (pid, pname, resp, ptype) in enumerate(ext_projects):
        fill = WHITE if pidx % 2 == 0 else LIGHT_ALT
        ws.cell(row, 1).value = pid
        ws.cell(row, 1).font = FONT_ID
        ws.cell(row, 2).value = pname
        ws.cell(row, 2).font = FONT_DATA
        ws.cell(row, 3).value = ptype
        ws.cell(row, 3).font = FONT_DATA
        for c in range(1, 12):
            ws.cell(row, c).fill = fill
            ws.cell(row, c).border = THIN_BORDER
            ws.cell(row, c).alignment = ALIGN_L if c >= 5 else ALIGN_C
        ws.row_dimensions[row].height = 50
        row += 1

    # Separator
    sep_row = row
    ws.row_dimensions[sep_row].height = 6
    row += 1

    # --- SECTION: INTERNE ---
    ws.merge_cells(f'A{row}:K{row}')
    ws.cell(row, 1).value = "INVESTISSEMENTS INTERNES"
    ws.cell(row, 1).font = Font(name='Arial', size=10, bold=True, color='E65100')
    ws.cell(row, 1).fill = SECTION_INT
    ws.cell(row, 1).alignment = ALIGN_L
    for col in range(2, 12):
        ws.cell(row, col).fill = SECTION_INT
    row += 1

    # Headers for interne
    for i, (col_letter, hdr_text, width) in enumerate(HEADERS):
        cell = ws.cell(row, i + 1)
        cell.value = hdr_text
        cell.font = FONT_HDR
        cell.fill = GRAY_HDR
        cell.alignment = ALIGN_C
        cell.border = THIN_BORDER
    row += 1

    # Interne projects
    int_projects = [p for p in PROJECTS if p[3] == "Interne"]
    for pidx, (pid, pname, resp, ptype) in enumerate(int_projects):
        fill = WHITE if pidx % 2 == 0 else LIGHT_ALT
        ws.cell(row, 1).value = pid
        ws.cell(row, 1).font = FONT_ID
        ws.cell(row, 2).value = pname
        ws.cell(row, 2).font = FONT_DATA
        ws.cell(row, 3).value = ptype
        ws.cell(row, 3).font = FONT_DATA
        for c in range(1, 12):
            ws.cell(row, c).fill = fill
            ws.cell(row, c).border = THIN_BORDER
            ws.cell(row, c).alignment = ALIGN_L if c >= 5 else ALIGN_C
        ws.row_dimensions[row].height = 50
        row += 1

    # Data validation: Statut
    statut_dv = DataValidation(
        type="list",
        formula1='"En cours,Termine,Suspendu,Annule,En attente"',
        allow_blank=True
    )
    statut_dv.error = "Choisir un statut valide"
    statut_dv.prompt = "Statut du projet"
    ws.add_data_validation(statut_dv)
    # Apply to all statut cells (col E, rows 6 to last)
    statut_dv.add(f'D6:D{row - 1}')

    # Type validation
    type_dv = DataValidation(
        type="list",
        formula1='"Externe,Interne"',
        allow_blank=True
    )
    ws.add_data_validation(type_dv)
    type_dv.add(f'C6:C{row - 1}')

    # Print / freeze
    ws.freeze_panes = 'A6'
    ws.page_setup.orientation = 'landscape'

# --- Config sheet ---
cfg = wb.create_sheet("Config")
cfg['A1'] = "Parametre"
cfg['B1'] = "Valeur"
cfg['A1'].font = Font(name='Arial', bold=True)
cfg['B1'].font = Font(name='Arial', bold=True)
cfg['A2'] = "Annee"
cfg['B2'] = 2026
cfg['A3'] = "Pole"
cfg['B3'] = "Investments"
cfg['A4'] = "Derniere_Semaine"
cfg['B4'] = "S11"
cfg['A5'] = "Responsable_Principal"
cfg['B5'] = "Imtiaz"
cfg.column_dimensions['A'].width = 25
cfg.column_dimensions['B'].width = 20

wb.save(OUT)
print(f"Created: {OUT}")
print(f"Sheets: {len(wb.sheetnames)} ({wb.sheetnames[0]} to {wb.sheetnames[-2]} + Config)")
