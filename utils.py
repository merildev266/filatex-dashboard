"""
utils.py — Shared utility functions for Excel parsing across all parsers.

These are pure functions with no side effects — they never read files, write files,
or make network calls. Import freely from any parser module.
"""
import math
import re
import unicodedata
from datetime import date, datetime
from typing import Optional, Union

from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Type aliases
# ---------------------------------------------------------------------------

CellValue = Union[str, int, float, datetime, date, None]


# ---------------------------------------------------------------------------
# Safe scalar converters
# ---------------------------------------------------------------------------

def sf(val: CellValue, default: float = 0.0) -> float:
    """
    Safe float conversion.

    Handles None, empty strings, non-numeric text, and NaN/Inf values.

    Args:
        val:     Raw cell value from openpyxl.
        default: Value to return when conversion fails (default 0.0).

    Returns:
        A finite float, or *default* if conversion is not possible.

    Examples:
        sf(None)        → 0.0
        sf("")          → 0.0
        sf("1,234.5")   → 0.0   (commas not handled — use ss() + strip first)
        sf(3.14)        → 3.14
        sf(float("nan"))→ 0.0
    """
    if val is None:
        return default
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


def sd(val: CellValue) -> Optional[str]:
    """
    Safe date conversion to ISO 8601 string ('YYYY-MM-DD').

    Accepts Python date/datetime objects (from openpyxl) or strings that
    already look like 'YYYY-MM-DD…'.

    Args:
        val: Raw cell value from openpyxl.

    Returns:
        'YYYY-MM-DD' string, or None if the value cannot be interpreted as a date.

    Examples:
        sd(datetime(2026, 3, 22))  → '2026-03-22'
        sd(date(2026, 3, 22))      → '2026-03-22'
        sd('2026-03-22T10:00:00')  → '2026-03-22'
        sd(None)                   → None
        sd('not a date')           → None
    """
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Accept 'YYYY-MM-DD' prefix (covers ISO 8601 datetime strings too)
    if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-":
        return s[:10]
    return None


def ss(val: CellValue) -> str:
    """
    Safe string conversion with whitespace stripping.

    Args:
        val: Raw cell value from openpyxl.

    Returns:
        Stripped string, or '' if val is None.

    Examples:
        ss(None)       → ''
        ss('  foo  ')  → 'foo'
        ss(42)         → '42'
    """
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Project ID helpers
# ---------------------------------------------------------------------------

def resolve_id(name: CellValue) -> Optional[str]:
    """
    Normalize a project name into a stable ASCII slug for cross-file matching.

    Steps:
      1. Unicode NFD decomposition → strip combining marks (accents).
      2. Lowercase.
      3. Replace runs of non-alphanumeric characters with underscore.
      4. Strip leading/trailing underscores.

    Args:
        name: Raw project name (cell value).

    Returns:
        ASCII slug string, or None if the result is empty.

    Examples:
        resolve_id('Énergie Solaire')  → 'energie_solaire'
        resolve_id('HFO — Tamatave')   → 'hfo_tamatave'
        resolve_id(None)               → None
    """
    if name is None:
        return None
    raw = str(name).strip()
    if not raw:
        return None
    # Strip diacritics
    normalized = unicodedata.normalize("NFD", raw.lower())
    ascii_only = "".join(c for c in normalized if unicodedata.category(c) != "Mn")
    # Replace non-alphanumeric runs with underscore
    slug = re.sub(r"[^a-z0-9]+", "_", ascii_only).strip("_")
    return slug or None


# ---------------------------------------------------------------------------
# Weekly workbook helpers
# ---------------------------------------------------------------------------

def find_latest_s_sheet(
    wb,
    min_col: int = 6,
    max_col: int = 15,
    data_start_row: int = 9,
) -> Optional[str]:
    """
    Find the highest-numbered 'S##' sheet in a weekly Excel workbook.

    Strategy:
      1. Check wb['Config']['B4'] — if it contains a valid 'S##' reference, use it.
      2. Otherwise scan all S##-named sheets and pick the one with the highest
         number that contains at least one non-empty cell in the data range.

    Args:
        wb:            openpyxl Workbook (read-only or normal).
        min_col:       First data column to check for non-empty cells (1-based).
        max_col:       Last data column to check (1-based).
        data_start_row: First row that should contain project data (1-based).

    Returns:
        Sheet name like 'S11', or None if no valid sheet is found.

    Examples:
        find_latest_s_sheet(wb)  → 'S11'
    """
    # 1. Check Config sheet
    if "Config" in wb.sheetnames:
        config_val = wb["Config"]["B4"].value
        if config_val and re.match(r"^S\d{2}$", str(config_val).strip()):
            return str(config_val).strip()

    # 2. Scan all S## sheets for data
    best_num: Optional[int] = None
    for sheet_name in wb.sheetnames:
        match = re.match(r"^S(\d{2})$", sheet_name)
        if not match:
            continue
        ws: Worksheet = wb[sheet_name]
        sheet_num = int(match.group(1))
        # Check whether any data rows contain non-empty cells
        end_row = min(ws.max_row + 1, data_start_row + 25)
        has_data = False
        for row in range(data_start_row, end_row):
            if any(
                ws.cell(row, col).value not in (None, "")
                for col in range(min_col, max_col + 1)
            ):
                has_data = True
                break
        if has_data and (best_num is None or sheet_num > best_num):
            best_num = sheet_num

    return f"S{best_num:02d}" if best_num is not None else None
