"""
sharepoint_client.py — Microsoft Graph API client for OneDrive/SharePoint access.

Authentication: Client Credentials flow via MSAL (no user login required).
All paths are relative to DATA_BASE_PATH configured in config.py.

Data safety: get_workbook() opens files read-only by default.
put_workbook() is ONLY used for the two pre-established comment write-back columns.
"""
import io
import logging
import os
import re
from typing import Optional

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

import config  # ensures load_dotenv() runs before we read env vars

log = logging.getLogger(__name__)

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# Module-level MSAL app — created once and reused (it maintains its own token cache)
_msal_app: Optional[ConfidentialClientApplication] = None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _get_msal_app() -> ConfidentialClientApplication:
    """Return the singleton MSAL ConfidentialClientApplication, creating it if needed."""
    global _msal_app
    if _msal_app is None:
        tenant_id = os.environ["AZURE_TENANT_ID"]
        client_id = os.environ["AZURE_CLIENT_ID"]
        client_secret = os.environ["AZURE_CLIENT_SECRET"]
        _msal_app = ConfidentialClientApplication(
            client_id=client_id,
            client_credential=client_secret,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
        )
        log.debug("MSAL app initialised (tenant=%s, client=%s)", tenant_id, client_id)
    return _msal_app


def get_token() -> str:
    """
    Acquire a Graph API access token with automatic cache refresh.

    MSAL's acquire_token_silent() returns a cached token if it is still valid,
    and falls back to acquire_token_for_client() (a network round-trip) only when
    the cached token has expired.

    Returns:
        A valid Bearer token string.

    Raises:
        RuntimeError: If authentication fails for any reason.
    """
    app = _get_msal_app()
    result = app.acquire_token_silent(GRAPH_SCOPES, account=None)
    if not result:
        log.debug("Token cache miss — acquiring new token from Azure AD")
        result = app.acquire_token_for_client(GRAPH_SCOPES)

    if "access_token" not in result:
        error = result.get("error", "unknown_error")
        description = result.get("error_description", "No description")
        raise RuntimeError(f"MSAL auth failed [{error}]: {description}")

    log.debug("Token acquired (expires_in=%s)", result.get("expires_in"))
    return result["access_token"]


def _auth_headers() -> dict[str, str]:
    """Return HTTP headers with a fresh Bearer token."""
    return {"Authorization": f"Bearer {get_token()}"}


def _drive_item_url(rel_path: str) -> str:
    """
    Build the Graph API URL for a drive item identified by its relative path.

    The full path is: DATA_BASE_PATH / rel_path
    Result is URL-encoded with forward slashes preserved.

    Args:
        rel_path: Path relative to DATA_BASE_PATH (e.g. '01_Energy/HFO/Tamatave/Tamatave_2026_01.xlsx')

    Returns:
        A complete Graph API URL for the drive item (without trailing action).
    """
    drive_id = os.environ["ONEDRIVE_DRIVE_ID"]
    base = config.DATA_BASE_PATH.rstrip("/")
    full_path = f"{base}/{rel_path.lstrip('/')}" if base else rel_path.lstrip("/")
    encoded = requests.utils.quote(full_path, safe="/")
    return f"{GRAPH_BASE}/drives/{drive_id}/root:/{encoded}"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_workbook(rel_path: str, read_only: bool = True) -> Workbook:
    """
    Download an Excel file from OneDrive and return an openpyxl Workbook.

    The file is never written to disk — it is streamed into memory via BytesIO.
    Formula results are read (data_only=True), not formula strings.

    Args:
        rel_path:  Path relative to DATA_BASE_PATH.
        read_only: Open in read-only mode (default True). Set False only when
                   you need to modify the workbook before calling put_workbook().

    Returns:
        An openpyxl Workbook object.

    Raises:
        requests.HTTPError: If the Graph API call fails (e.g. 404 not found).
        RuntimeError:       If authentication fails.
    """
    url = _drive_item_url(rel_path) + ":/content"
    log.info("Downloading: %s", rel_path)
    resp = requests.get(url, headers=_auth_headers(), timeout=30)
    resp.raise_for_status()
    wb = load_workbook(
        io.BytesIO(resp.content),
        read_only=read_only,
        data_only=True,
    )
    log.debug("Opened workbook '%s' — sheets: %s", rel_path, wb.sheetnames)
    return wb


def get_file_bytes(rel_path: str) -> bytes:
    """Download a file from OneDrive and return raw bytes (for pandas ExcelFile use)."""
    url = _drive_item_url(rel_path) + ":/content"
    log.info("Downloading bytes: %s", rel_path)
    resp = requests.get(url, headers=_auth_headers(), timeout=30)
    resp.raise_for_status()
    return resp.content


def put_workbook(wb: Workbook, rel_path: str) -> None:
    """
    Upload a modified workbook back to OneDrive.

    IMPORTANT: This function exists solely for writing comment columns
    (P/Q in ENR weekly, J/K in Investments weekly). Do NOT use it for
    any other purpose — Excel files are the immutable source of truth.

    Args:
        wb:       The modified openpyxl Workbook to upload.
        rel_path: Destination path relative to DATA_BASE_PATH.

    Raises:
        requests.HTTPError: If the upload fails.
        RuntimeError:       If authentication fails.
    """
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    url = _drive_item_url(rel_path) + ":/content"
    log.info("Uploading: %s (%d bytes)", rel_path, buf.getbuffer().nbytes)
    resp = requests.put(
        url,
        headers={**_auth_headers(), "Content-Type": content_type},
        data=buf.read(),
        timeout=60,
    )
    resp.raise_for_status()
    log.info("Upload complete: %s", rel_path)


def list_folder(rel_path: str) -> list[dict]:
    """
    List files and folders inside an OneDrive folder.

    Args:
        rel_path: Path relative to DATA_BASE_PATH.

    Returns:
        A list of item dicts, each containing at minimum:
          - name (str)
          - id (str)
          - lastModifiedDateTime (str, ISO 8601)
          - size (int, bytes) — for files only
          - folder / file sub-dict
    """
    url = _drive_item_url(rel_path) + ":/children"
    log.debug("Listing folder: %s", rel_path)
    resp = requests.get(url, headers=_auth_headers(), timeout=15)
    resp.raise_for_status()
    items: list[dict] = resp.json().get("value", [])
    log.debug("Found %d items in '%s'", len(items), rel_path)
    return items


def find_latest_monthly_file(
    folder: str,
    prefix: str,
    year: int,
) -> Optional[str]:
    """
    Find the most recent monthly Excel file matching '<prefix>_<year>_<NN>.xlsx'.

    Scans the given folder and returns the relative path (within DATA_BASE_PATH)
    of the file with the highest month number for the specified year.

    Example:
        find_latest_monthly_file('01_Energy/HFO/Tamatave', 'Tamatave', 2026)
        → '01_Energy/HFO/Tamatave/Tamatave_2026_03.xlsx'

    Args:
        folder: Folder path relative to DATA_BASE_PATH.
        prefix: File name prefix (e.g. 'Tamatave', 'Diego').
        year:   The four-digit year to search for.

    Returns:
        Relative path of the latest monthly file, or None if none found.
    """
    pattern = re.compile(
        rf"^{re.escape(prefix)}_{year}_(\d{{2}})\.xlsx$",
        re.IGNORECASE,
    )
    items = list_folder(folder)
    best_month: Optional[int] = None
    best_name: Optional[str] = None

    for item in items:
        name: str = item.get("name", "")
        m = pattern.match(name)
        if m:
            month = int(m.group(1))
            if best_month is None or month > best_month:
                best_month = month
                best_name = name

    if best_name is None:
        log.warning(
            "No monthly file found for prefix='%s', year=%d in '%s'",
            prefix, year, folder,
        )
        return None

    rel_path = f"{folder.rstrip('/')}/{best_name}"
    log.info(
        "Latest monthly file for '%s' %d: %s (month %02d)",
        prefix, year, rel_path, best_month,
    )
    return rel_path


def find_monthly_file(
    folder: str,
    prefix: str,
    year: int,
    month: int,
) -> Optional[str]:
    """
    Return the relative path for a specific month's Excel file, or None if not found.

    Does a lightweight metadata GET (no content download) to confirm the file exists.

    Args:
        folder: Folder path relative to DATA_BASE_PATH.
        prefix: File name prefix (e.g. 'Diego').
        year:   Four-digit year.
        month:  Month number (1–12).

    Returns:
        Relative path like 'folder/Prefix_year_MM.xlsx', or None.
    """
    target_name = f"{prefix}_{year}_{month:02d}.xlsx"
    rel_path = f"{folder.rstrip('/')}/{target_name}"
    url = _drive_item_url(rel_path)
    try:
        resp = requests.get(url, headers=_auth_headers(), timeout=15)
        if resp.status_code == 200:
            log.debug("find_monthly_file: found %s", rel_path)
            return rel_path
        log.debug("find_monthly_file: %s → HTTP %s", target_name, resp.status_code)
        return None
    except Exception as exc:
        log.warning("find_monthly_file: error checking %s: %s", target_name, exc)
        return None


def list_available_months(
    folder: str,
    prefix: str,
    year: int,
) -> list[int]:
    """
    Return a sorted list of month numbers that have Excel files in an OneDrive folder.

    Args:
        folder: Folder path relative to DATA_BASE_PATH.
        prefix: File name prefix (e.g. 'Diego').
        year:   Four-digit year.

    Returns:
        Sorted list of month numbers present, e.g. [1, 2, 3].
    """
    pattern = re.compile(
        rf"^{re.escape(prefix)}_{year}_(\d{{2}})\.xlsx$",
        re.IGNORECASE,
    )
    try:
        items = list_folder(folder)
    except Exception as exc:
        log.warning("list_available_months: cannot list %s: %s", folder, exc)
        return []
    months = []
    for item in items:
        m = pattern.match(item.get("name", ""))
        if m:
            months.append(int(m.group(1)))
    return sorted(months)
