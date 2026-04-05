#!/usr/bin/env python3
"""
Génère com_reporting_data.js depuis COM_Reporting.xlsx
Structure identique à props_data_dev_full.js (DEV)
Usage: python generate_com_reporting.py
"""
import openpyxl
import json
import os

EXCEL_PATH = os.path.join(
    os.path.expanduser('~'),
    'OneDrive - GROUPE FILATEX',
    'Bureau',
    'Fichiers de DOSSIER DASHBOARD - Data_Dashbords',
    '02_Properties', 'Reporting', 'COM_Reporting.xlsx'
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'js', 'com_reporting_data.js')


def read_sheet(wb, sheet_name, has_phase=True):
    """Lit une feuille du COM_Reporting et retourne la structure par item."""
    ws = wb[sheet_name]
    items = {}  # name -> [{ week, phase, avancement, comment }]

    # Colonnes:
    # Vente Projet/Terrain: A=Semaine, B=Projet/Terrain, C=Phase, D=Avancement, E=Commentaire
    # Location: A=Semaine, B=Bien, C=Avancement, D=Commentaire
    current_week = ''
    for row in ws.iter_rows(min_row=5, values_only=False):
        # Lire semaine (colonne A, peut être merged/vide)
        week_val = row[0].value
        if week_val and isinstance(week_val, str) and week_val.startswith('S'):
            current_week = week_val.strip()

        if not current_week:
            continue

        # Lire nom item (colonne B)
        name = row[1].value
        if not name or not isinstance(name, str) or name.strip() == '':
            continue
        name = name.strip()

        if has_phase:
            phase = str(row[2].value or '').strip()
            avancement = str(row[3].value or '').strip()
            comment = str(row[4].value or '').strip()
        else:
            phase = ''
            avancement = str(row[2].value or '').strip()
            comment = str(row[3].value or '').strip()

        # Ne garder que les semaines avec du contenu
        if not phase and not avancement and not comment:
            continue

        if name not in items:
            items[name] = []
        items[name].append({
            'week': current_week,
            'phase': phase,
            'avancement': avancement,
            'comment': comment
        })

    return items


def build_js():
    if not os.path.exists(EXCEL_PATH):
        print(f'ERREUR: Fichier non trouvé: {EXCEL_PATH}')
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f'Lecture de {EXCEL_PATH}')
    print(f'Feuilles: {wb.sheetnames}')

    result = {}

    # Vente Projet (avec Phase)
    if 'Vente Projet' in wb.sheetnames:
        result['venteProjet'] = []
        items = read_sheet(wb, 'Vente Projet', has_phase=True)
        for name, history in items.items():
            result['venteProjet'].append({'name': name, 'history': history})
        print(f'  Vente Projet: {len(result["venteProjet"])} projets')

    # Vente Terrain (avec Phase)
    if 'Vente Terrain' in wb.sheetnames:
        result['venteTerrain'] = []
        items = read_sheet(wb, 'Vente Terrain', has_phase=True)
        for name, history in items.items():
            result['venteTerrain'].append({'name': name, 'history': history})
        print(f'  Vente Terrain: {len(result["venteTerrain"])} terrains')

    # Location (sans Phase)
    if 'Location' in wb.sheetnames:
        result['location'] = []
        items = read_sheet(wb, 'Location', has_phase=False)
        for name, history in items.items():
            result['location'].append({'name': name, 'history': history})
        print(f'  Location: {len(result["location"])} biens')

    # Générer le JS
    js = '/* COM Reporting Data - auto-generated from COM_Reporting.xlsx */\n'
    js += '/* Ne pas éditer à la main — relancer generate_com_reporting.py */\n\n'

    for key, label in [('venteProjet', 'venteProjet'), ('venteTerrain', 'venteTerrain'), ('location', 'location')]:
        data = result.get(key, [])
        js += f'var comReport_{label} = [\n'
        for item in data:
            js += '  {\n'
            js += f'    name: {json.dumps(item["name"], ensure_ascii=False)},\n'
            js += '    history: [\n'
            for h in item['history']:
                js += '      {week:%s, phase:%s, avancement:%s, comment:%s},\n' % (
                    json.dumps(h['week'], ensure_ascii=False),
                    json.dumps(h['phase'], ensure_ascii=False),
                    json.dumps(h['avancement'], ensure_ascii=False),
                    json.dumps(h['comment'], ensure_ascii=False)
                )
            js += '    ]\n'
            js += '  },\n'
        js += '];\n\n'

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js)

    print(f'\nGénéré: {OUTPUT_PATH}')


if __name__ == '__main__':
    build_js()
