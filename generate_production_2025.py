"""
Pipeline — Production 2025 par site (4 sites HFO).

Lit les fichiers mensuels Excel de chaque site dans OneDrive :
    01_Energy/1.Production/3.Production_2025/{Antsirabe,Majunga,Tamatave,Tulear}/

Normalise la production journaliere (Net, kWh -> MWh) et ecrit
`frontend/src/data/production_2025.js` avec pour chaque site :
  - monthly : [{ m: 1..12, prod_mwh }]
  - daily   : [{ date: 'YYYY-MM-DD', prod_mwh }]

Note : Majunga est conserve avec zeros (moteur arrete janvier 2025).
Solaire de Tulear NON inclus (reste dans la section EnR du dashboard).
"""
import json
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(
    r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau"
    r"\Fichiers de DOSSIER DASHBOARD - Data_Dashbords"
    r"\01_Energy\1.Production\3.Production_2025"
)

OUT_PATH = Path("frontend/src/data/production_2025.js")


def _parse_date(val, expected_month=None):
    """Return 'YYYY-MM-DD'.

    Strategie : l'annee saisie par les equipes n'est pas toujours 2025 (oubli
    frequent sur les templates). On force l'annee a 2025 et on accepte la ligne
    si :
      - `expected_month` est fourni -> le mois du fichier doit correspondre au
        mois de la date saisie (fiabilite jour/mois > annee),
      - sinon on accepte tout datetime.
    """
    if not isinstance(val, datetime):
        return None
    if expected_month is not None and val.month != expected_month:
        return None
    return f"2025-{val.month:02d}-{val.day:02d}"


def _month_from_filename(xlsx_path):
    """Extrait le mois (1..12) du nom de fichier type Site_2025_MM.xlsx."""
    stem = xlsx_path.stem  # ex. 'Tamatave_2025_02'
    try:
        return int(stem.split("_")[-1])
    except (ValueError, IndexError):
        return None


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def load_antsirabe(xlsx_path):
    """Antsirabe : 2 formats selon la periode.

    - Jan-Mai 2025 : feuille `COVER ` — col B=Date, col I=Total Generation (kWh)
    - Jun-Dec 2025 : site agrandi (3 moteurs). Feuille `SFOC SLOC` —
      col A=Date, cols B+C+D = BDG01 + BDG02 + BDG03 production (kWh)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    out = []
    names = set(wb.sheetnames)
    exp_m = _month_from_filename(xlsx_path)
    if "COVER " in names:
        ws = wb["COVER "]
        for row in ws.iter_rows(min_row=4, values_only=True):
            d = _parse_date(row[1] if len(row) > 1 else None, exp_m)
            if not d:
                continue
            net_kwh = _num(row[8] if len(row) > 8 else 0)
            out.append({"date": d, "net_mwh": net_kwh / 1000.0})
    elif "SFOC SLOC" in names:
        ws = wb["SFOC SLOC"]
        # Trouver la ligne d'en-tete contenant >=2 colonnes moteurs consecutives
        # nommees "BDG01/02/03" (ancien) ou "CUMMINS 1..4 (kwh)" (nouveau).
        data_cols = []
        for hdr_row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            best = []
            current = []
            for i, cell in enumerate(hdr_row):
                s = str(cell).lower() if cell is not None else ""
                is_eng = ("bdg0" in s or "bdg 0" in s
                          or ("cummins" in s and "kwh" in s))
                if is_eng:
                    if current and i == current[-1] + 1:
                        current.append(i)
                    else:
                        current = [i]
                    if len(current) > len(best):
                        best = list(current)
                else:
                    current = []
            if len(best) >= 2:
                data_cols = best
                break
        if not data_cols:
            wb.close()
            return out
        date_col = data_cols[0] - 1
        if date_col < 0:
            wb.close()
            return out
        for row in ws.iter_rows(min_row=5, values_only=True):
            d = _parse_date(row[date_col] if len(row) > date_col else None, exp_m)
            if not d:
                continue
            total = sum(_num(row[c]) for c in data_cols if c < len(row))
            out.append({"date": d, "net_mwh": total / 1000.0})
    wb.close()
    return out


def load_majunga(xlsx_path):
    """RME [Month] 2025 sheet — col A=Date, col E (idx 4)=Net Export (kWh)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    exp_m = _month_from_filename(xlsx_path)
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        d = _parse_date(row[0] if len(row) > 0 else None, exp_m)
        if not d:
            continue
        net_kwh = _num(row[4] if len(row) > 4 else 0)
        out.append({"date": d, "net_mwh": net_kwh / 1000.0})
    wb.close()
    return out


def load_tamatave(xlsx_path):
    """Daily Data sheet — col A=Date, col C (idx 2)=Net (kWh, header dit MWh mais valeurs kWh).

    Note: les equipes oublient parfois d'ajuster l'annee des dates (2023 au
    lieu de 2025). Le mois du fichier fait foi via `_parse_date`.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Daily Data"]
    exp_m = _month_from_filename(xlsx_path)
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        d = _parse_date(row[0] if len(row) > 0 else None, exp_m)
        if not d:
            continue
        net_kwh = _num(row[2] if len(row) > 2 else 0)
        out.append({"date": d, "net_mwh": net_kwh / 1000.0})
    wb.close()
    return out


def load_tulear(xlsx_path):
    """DAILY DATA sheet — col A=Date, col C (idx 2)=Net (kWh). HFO uniquement (DG)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["DAILY DATA"]
    exp_m = _month_from_filename(xlsx_path)
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        d = _parse_date(row[0] if len(row) > 0 else None, exp_m)
        if not d:
            continue
        net_kwh = _num(row[2] if len(row) > 2 else 0)
        out.append({"date": d, "net_mwh": net_kwh / 1000.0})
    wb.close()
    return out


LOADERS = {
    "antsirabe": ("Antsirabe", load_antsirabe),
    "majunga":   ("Majunga",   load_majunga),
    "tamatave":  ("Tamatave",  load_tamatave),
    "tulear":    ("Tulear",    load_tulear),
}


def build_site(site_key, folder_name, loader):
    """Produit une structure COMPACTE :
      - monthly : list[12] de float  (MWh, index 0..11)
      - daily   : dict { "01".."12" -> list[n] de float } (n = nb jours du mois)

    Format minimal pour limiter la taille du bundle JS (2025 = donnees figees).
    """
    import calendar
    folder = BASE / folder_name
    empty = {
        "monthly": [0.0] * 12,
        "daily":   {f"{m:02d}": [0.0] * calendar.monthrange(2025, m)[1] for m in range(1, 13)},
    }
    if not folder.exists():
        print(f"  [skip] {folder} introuvable")
        return empty

    by_date = {}
    for xlsx in sorted(folder.glob("*.xlsx")):
        try:
            rows = loader(xlsx)
            total = 0.0
            for r in rows:
                by_date[r["date"]] = r["net_mwh"]
                total += r["net_mwh"]
            print(f"  {xlsx.name}: {len(rows)} jours, total {total:.1f} MWh")
        except Exception as e:
            print(f"  [err] {xlsx.name}: {e}")

    daily_map = {f"{m:02d}": [0.0] * calendar.monthrange(2025, m)[1] for m in range(1, 13)}
    monthly = [0.0] * 12
    for date_str, v in by_date.items():
        mm = date_str[5:7]
        dd = int(date_str[8:10])
        if 1 <= dd <= len(daily_map[mm]):
            daily_map[mm][dd - 1] = round(v, 1)
            monthly[int(mm) - 1] += v
    monthly = [round(x, 1) for x in monthly]
    return {"monthly": monthly, "daily": daily_map}


def main():
    all_data = {}
    for key, (folder_name, loader) in LOADERS.items():
        print(f"\n=== {key.upper()} ===")
        all_data[key] = build_site(key, folder_name, loader)
        total_year = sum(all_data[key]["monthly"])
        n_jours = sum(1 for arr in all_data[key]["daily"].values() for v in arr if v > 0)
        print(f"  TOTAL 2025 : {total_year:,.0f} MWh ({n_jours} jours avec donnees)")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        f.write("// Auto-generated by generate_production_2025.py - do not edit\n")
        f.write("// Source : OneDrive .../01_Energy/1.Production/3.Production_2025/\n")
        f.write("// Unite : MWh (Net). Majunga = 0 (moteur arrete). Tulear HFO uniquement.\n")
        f.write("// Format compact (donnees figees 2025) :\n")
        f.write("//   monthly : number[12]   (index 0=Jan .. 11=Dec)\n")
        f.write("//   daily   : { '01'..'12' : number[n_jours_du_mois] }  (index 0 = 1er du mois)\n\n")
        f.write("export const PRODUCTION_2025 = ")
        f.write(json.dumps(all_data, separators=(",", ":"), ensure_ascii=False))
        f.write(";\n")
    print(f"\n[OK] Ecrit : {OUT_PATH}")


if __name__ == "__main__":
    main()
