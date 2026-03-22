"""
Flask application serving the Filatex PMO Dashboard.
Reads Tamatave xlsx data and provides it via JSON API.
Supports DG comment writing back to Excel files.
Phase 2: SharePoint API endpoints registered via Blueprint.
Phase 4: Structured logging, background cache pre-warm, error hardening.
"""
import logging
import os
import re
import openpyxl
from flask import Flask, jsonify, request, send_from_directory
from data_loader import build_tamatave_data
import config
import cache
import api_data

# ── Configure logging before anything else ──────────────────────────────────
logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL.upper(), logging.INFO),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
_log = logging.getLogger(__name__)
_log.info("Starting Filatex PMO Dashboard (log_level=%s)", config.LOG_LEVEL)

app = Flask(__name__, static_folder=".", static_url_path="")

# ── Phase 2: Register SharePoint API Blueprint ──
app.register_blueprint(api_data.api)

# ── Background cache refresh (only when SharePoint is configured) ──
_missing = config.validate()
if not _missing:
    # Register all cache keys for pre-warm background refresh
    def _make_hfo_loader():
        from parsers import hfo
        return hfo.build_all_sites

    def _make_enr_sites_loader():
        from parsers import enr_sites
        return enr_sites.build

    def _make_hfo_proj_loader():
        from parsers import hfo_projects
        return hfo_projects.build

    def _make_enr_proj_loader():
        from parsers import enr_projects
        return enr_projects.build

    def _make_capex_loader():
        from parsers import capex_parser
        return capex_parser.build

    def _make_reporting_loader():
        from parsers import reporting
        return reporting.build

    cache.register("hfo_sites", _make_hfo_loader(), config.TTL_HFO_SITES)
    cache.register("enr_sites", _make_enr_sites_loader(), config.TTL_ENR_SITES)
    cache.register("hfo_projects", _make_hfo_proj_loader(), config.TTL_HFO_PROJECTS)
    cache.register("enr_projects", _make_enr_proj_loader(), config.TTL_ENR_PROJECTS)
    cache.register("capex", _make_capex_loader(), config.TTL_CAPEX)
    cache.register("enr_reporting", _make_reporting_loader(), config.TTL_REPORTING)
    # Background worker performs an immediate first pass on startup (pre-warms
    # all registered keys), then continues refreshing on the normal interval.
    cache.start_background_refresh()
    _log.info("SharePoint configured — background cache refresh started")
else:
    _log.warning(
        "SharePoint not configured (missing: %s) — background refresh disabled",
        ", ".join(_missing),
    )

BASE_ENR = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - GROUPE FILATEX",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "01_Energy", "Projet", "EnR"
)
BASE_INV = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - GROUPE FILATEX",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "03_ Investments", "Reporting"
)


def _find_latest_sheet(wb):
    """Find the latest filled S## sheet in a workbook.
    Checks user-editable columns (6+) to distinguish filled vs template-only sheets.
    """
    # First check Config sheet for explicit setting
    if "Config" in wb.sheetnames:
        cfg_val = wb["Config"]["B4"].value
        if cfg_val and re.match(r"^S\d{2}$", str(cfg_val)):
            return str(cfg_val)

    latest = None
    for sn in wb.sheetnames:
        m = re.match(r"^S(\d{2})$", sn)
        if not m:
            continue
        ws = wb[sn]
        # Check editable columns (6+: Phase, Avancement, etc.) for real user data
        has_data = False
        for r in range(9, min(ws.max_row + 1, 35)):
            for c in range(6, min(ws.max_column + 1, 18)):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    has_data = True
                    break
            if has_data:
                break
        if has_data:
            num = int(m.group(1))
            if latest is None or num > latest:
                latest = num
    return f"S{latest:02d}" if latest else None


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/api/tamatave")
def api_tamatave():
    data = build_tamatave_data()
    if data is None:
        return jsonify({"error": "No data found"}), 404
    return jsonify(data)


@app.route("/api/comment/enr", methods=["POST"])
def save_enr_comment():
    """Save a DG comment and/or response for an EnR project into the weekly Excel file."""
    data = request.get_json()
    if not data or "projectId" not in data:
        return jsonify({"error": "projectId required"}), 400

    pid = data["projectId"]
    comment = data.get("comment")
    reponse = data.get("reponse")
    if comment is None and reponse is None:
        return jsonify({"error": "comment or reponse required"}), 400

    # Find the weekly file
    path = os.path.join(BASE_ENR, "Weekly_Report", "Weekly_EnR_Avancement.xlsx")
    if not os.path.exists(path):
        path = os.path.join(BASE_ENR, "Weekly_EnR_Avancement.xlsx")
    if not os.path.exists(path):
        return jsonify({"error": "Weekly EnR file not found"}), 404

    try:
        wb = openpyxl.load_workbook(path)
        sheet_name = _find_latest_sheet(wb)
        if not sheet_name:
            wb.close()
            return jsonify({"error": "No active weekly sheet found"}), 404

        ws = wb[sheet_name]
        # comDG = column 16 (P), Reponse = column 17 (Q)
        col_com = 16
        col_rep = 17
        updated = False
        for r in range(6, ws.max_row + 1):
            cell_pid = ws.cell(r, 1).value
            if cell_pid and str(cell_pid).strip() == pid:
                if comment is not None:
                    ws.cell(r, col_com).value = comment.strip()
                if reponse is not None:
                    ws.cell(r, col_rep).value = reponse.strip()
                updated = True
                break

        if not updated:
            wb.close()
            return jsonify({"error": f"Project {pid} not found in {sheet_name}"}), 404

        wb.save(path)
        wb.close()
        return jsonify({"ok": True, "sheet": sheet_name, "project": pid})

    except PermissionError:
        return jsonify({"error": "Fichier Excel ouvert dans une autre application. Fermez-le et reessayez."}), 423
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/comment/investments", methods=["POST"])
def save_inv_comment():
    """Save a DG comment and/or response for an Investments project into the weekly Excel file."""
    data = request.get_json()
    if not data or "projectId" not in data:
        return jsonify({"error": "projectId required"}), 400

    pid = data["projectId"]
    comment = data.get("comment")
    reponse = data.get("reponse")
    if comment is None and reponse is None:
        return jsonify({"error": "comment or reponse required"}), 400

    # Find the weekly file
    path = os.path.join(BASE_INV, "Weekly_Investments_Avancement.xlsx")
    if not os.path.exists(path):
        path = os.path.join(BASE_INV, "Weekly_Investments_Avancement_v2.xlsx")
    if not os.path.exists(path):
        return jsonify({"error": "Weekly Investments file not found"}), 404

    try:
        wb = openpyxl.load_workbook(path)
        sheet_name = _find_latest_sheet(wb)
        if not sheet_name:
            wb.close()
            return jsonify({"error": "No active weekly sheet found"}), 404

        ws = wb[sheet_name]
        # commentaires DG = column 10 (J), Reponse = column 11 (K)
        col_com = 10
        col_rep = 11
        updated = False
        for r in range(6, ws.max_row + 1):
            cell_pid = ws.cell(r, 1).value
            if cell_pid and str(cell_pid).strip() == pid:
                if comment is not None:
                    ws.cell(r, col_com).value = comment.strip()
                if reponse is not None:
                    ws.cell(r, col_rep).value = reponse.strip()
                updated = True
                break

        if not updated:
            wb.close()
            return jsonify({"error": f"Project {pid} not found in {sheet_name}"}), 404

        wb.save(path)
        wb.close()
        return jsonify({"ok": True, "sheet": sheet_name, "project": pid})

    except PermissionError:
        return jsonify({"error": "Fichier Excel ouvert dans une autre application. Fermez-le et reessayez."}), 423
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    # Cloud Run sets PORT; fall back to 5000 for local dev
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
