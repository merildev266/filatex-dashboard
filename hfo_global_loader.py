"""
Parses Daily_Report_Global.xlsx.

Produces the 'truth' for:
  - contract / peak / MW / dispo per site (split ENELEC / VESTOP / LFO)
  - per-engine nominal / attendu / available / status / stop date / issues
  - overhaul plan (OVERHAUL sheet)
  - weekly planning per engine 2026 with color codes (HEBDO sheet)

All per-site outputs are keyed by the site keys used elsewhere:
  'tamatave', 'diego', 'tulear', 'majunga', 'antsirabe', 'fihaonana'

Call ``load_global()`` to get a single dict containing everything.
"""
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import hfo_config as cfg


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def _s(v):
    """Safe string: None/nan → ''."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def _f(v):
    """Safe float: returns None when not parseable."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _iso(v):
    """Safe ISO date string: returns None when not a date."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    try:
        return pd.Timestamp(v).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _site_key_from_label(label):
    """Map RECAP/OVERHAUL site labels to our internal site keys."""
    if not label:
        return None
    s = str(label).strip().lower()
    # strip diacritics lightly
    s = (s.replace("é", "e").replace("è", "e").replace("ê", "e")
           .replace("â", "a").replace("à", "a").replace("ô", "o"))
    if s.startswith("tamatave"):
        return "tamatave"
    if s.startswith("majunga"):
        return "majunga"
    if s.startswith("diego") or s.startswith("di"):
        return "diego"
    if s.startswith("tulear") or s.startswith("tul"):
        return "tulear"
    if s.startswith("antsirabe"):
        return "antsirabe"
    if s.startswith("fihaonana"):
        return "fihaonana"
    return None


# ══════════════════════════════════════════════════════════════
# OVERHAUL sheet
# ══════════════════════════════════════════════════════════════

def load_overhaul(xlsx_path):
    """
    Returns a flat list of overhaul plan rows:
        [{engine, society, site, pieces_dispo, pieces_attente,
          date_pieces_init, date_pieces_revu, glissement_jours,
          date_ovh_init_debut, date_ovh_init_fin, situation,
          date_ovh_revu_debut, date_ovh_revu_fin}, ...]
    """
    df = pd.read_excel(xlsx_path, sheet_name=cfg.SHEET_OVERHAUL, header=None)
    rows = []
    # Data starts at row index 3 (header rows 1-2)
    for i in range(3, df.shape[0]):
        engine = _s(df.iloc[i, 0])
        if not engine:
            continue
        rows.append({
            "engine":              engine,
            "society":             _s(df.iloc[i, 1]),
            "site":                _s(df.iloc[i, 2]),
            "siteKey":             _site_key_from_label(df.iloc[i, 2]),
            "piecesDispo":         _f(df.iloc[i, 3]),
            "piecesEnAttente":     _f(df.iloc[i, 4]),
            "datePiecesInit":      _iso(df.iloc[i, 5]),
            "datePiecesRevu":      _iso(df.iloc[i, 6]),
            "glissementJours":     _f(df.iloc[i, 7]),
            "dateOvhInitDebut":    _iso(df.iloc[i, 8]),
            "dateOvhInitFin":      _iso(df.iloc[i, 9]) if df.shape[1] > 9 else None,
            "situation":           _s(df.iloc[i, 10]) if df.shape[1] > 10 else "",
            "dateOvhRevuDebut":    _iso(df.iloc[i, 11]) if df.shape[1] > 11 else None,
            "dateOvhRevuFin":      _iso(df.iloc[i, 12]) if df.shape[1] > 12 else None,
        })
    return rows


# ══════════════════════════════════════════════════════════════
# RECAP sheet
# ══════════════════════════════════════════════════════════════

def load_recap(xlsx_path):
    """
    Returns a per-site dict with providers split:
        {
          tamatave: {
            providers: {
              enelec: {contrat, pointe, nbTotal, nbOperationnel, prodPlan, dispo,
                       dispoVsPlan, dispoVsContrat, dispoVsPeak, delestage},
              vestop: {nbTotal, nbOperationnel, dispo, ...},
              lfo:    {nbTotal, dispo, ...},
            },
            lubes: [{fournisseur, securiteLitres, stockLitres, jusquA}, ...],
          }, ...
        }
    """
    df = pd.read_excel(xlsx_path, sheet_name=cfg.SHEET_RECAP, header=None)
    out = {}
    i = 3  # first data row
    current_site_key = None

    def _pct(v):
        """RECAP stores pct as 0.xx floats; return integer percent."""
        f = _f(v)
        return None if f is None else round(f * 100, 1)

    while i < df.shape[0]:
        label = _s(df.iloc[i, 1])
        if not label or label.lower() == "total":
            i += 1
            continue

        low = label.lower()
        key = _site_key_from_label(label)

        # Determine provider from label suffix
        if "vestop" in low:
            provider = cfg.PROVIDER_VESTOP
        elif "lfo" in low:
            provider = cfg.PROVIDER_LFO
        else:
            provider = cfg.PROVIDER_ENELEC
            current_site_key = key

        sk = key or current_site_key
        if sk is None:
            i += 1
            continue

        entry = out.setdefault(sk, {"providers": {}, "lubes": []})
        entry["providers"][provider] = {
            "contrat":         _f(df.iloc[i, 2]),
            "pointe":          _f(df.iloc[i, 3]),
            "nbTotal":         _f(df.iloc[i, 4]),
            "nbOperationnel":  _f(df.iloc[i, 5]),
            "prodPlan":        _f(df.iloc[i, 6]),
            "dispo":           _f(df.iloc[i, 7]),
            "dispoVsPlan":     _pct(df.iloc[i, 8]),
            "dispoVsContrat":  _pct(df.iloc[i, 9]),
            "dispoVsPeak":     _pct(df.iloc[i, 10]),
            "delestage":       _s(df.iloc[i, 11]),
        }
        # Lubes block (cols 12-15) — several rows per site for Fournisseur/TOTAL etc.
        fourn = _s(df.iloc[i, 12])
        if fourn:
            entry["lubes"].append({
                "fournisseur":    fourn,
                "securiteLitres": _f(df.iloc[i, 13]),
                "stockLitres":    _f(df.iloc[i, 14]),
                "jusquA":         _iso(df.iloc[i, 15]),
            })
        # Rows without a label but with lube info: attach to current site
        i += 1
        while i < df.shape[0] and not _s(df.iloc[i, 1]):
            fourn2 = _s(df.iloc[i, 12])
            if fourn2:
                entry["lubes"].append({
                    "fournisseur":    fourn2,
                    "securiteLitres": _f(df.iloc[i, 13]),
                    "stockLitres":    _f(df.iloc[i, 14]),
                    "jusquA":         _iso(df.iloc[i, 15]),
                })
            i += 1

    return out


# ══════════════════════════════════════════════════════════════
# Situation moteur sheet
# ══════════════════════════════════════════════════════════════

SITE_HEADER_RE = re.compile(r"^(tamatave|majunga|di[eé]go|tul[eé]ar|antsirabe|fihaonana)\s*site", re.I)


def load_situation_moteurs(xlsx_path):
    """
    Returns per-site engine list:
        {
          tamatave: [
            {id, make, model, nominal, attendu, availableMw, availableRaw,
             statut, lubes, stopDate, issues, actions, dlRevu, fuel, provider}
          ],
          ...
        }

    - nominal = col "1.0"  (plate)
    - attendu = col "0.85" (contractual continuous limit)
    - availableMw = numeric "Available" if parseable, else None
    - availableRaw = original cell text (used as status label when non-numeric)
    - statut = " statuse" column (Working/Breakdown/Stand By/UM/Not exist/Installation/NEW/Under Project)
    """
    df = pd.read_excel(xlsx_path, sheet_name=cfg.SHEET_SITUATION, header=None)
    out = {}
    current_site = None

    for i in range(df.shape[0]):
        label = _s(df.iloc[i, 0])
        if not label:
            continue
        # Site header row
        m = SITE_HEADER_RE.match(label)
        if m:
            current_site = _site_key_from_label(label)
            if current_site:
                out.setdefault(current_site, [])
            continue
        # Skip header sub-row
        if label.lower() == "s.no.":
            continue
        # Skip Total row
        if label.lower() == "total":
            current_site = None  # end of site block
            continue
        # Engine row — need a site context AND a valid engine name in col B
        if not current_site:
            continue
        engine = _s(df.iloc[i, 1])
        if not engine or engine.lower() == "vestop":
            # Vestop engines under Majunga use "Vestop" literally — keep them with Make+Model
            if not engine:
                continue
            pass

        make = _s(df.iloc[i, 2])
        model = _s(df.iloc[i, 3])
        nominal = _f(df.iloc[i, 4])
        attendu = _f(df.iloc[i, 5])
        avail_raw = df.iloc[i, 6]
        avail_num = _f(avail_raw)
        statut = _s(df.iloc[i, 7])
        lubes = _s(df.iloc[i, 8])
        stop_date_raw = df.iloc[i, 9]
        issues = _s(df.iloc[i, 10]) if df.shape[1] > 10 else ""
        actions = _s(df.iloc[i, 11]) if df.shape[1] > 11 else ""
        dl_revu = _iso(df.iloc[i, 12]) if df.shape[1] > 12 else None

        # Classify provider + fuel
        # VESTOP: Majunga rows literally "Vestop" (6-7), plus ADG14 Deutz (Tamatave),
        #         Antsirabe BDG1/2/3 Pielstick + Wartsila R32 (all VESTOP), all Cummins on Antsirabe
        # LFO: Diego Cummins (DCum*)
        eu = engine.upper()
        provider = cfg.PROVIDER_ENELEC
        fuel = "HFO"
        if current_site == "antsirabe":
            # All Antsirabe engines are VESTOP
            provider = cfg.PROVIDER_VESTOP
            if "CUM" in eu:
                fuel = "LFO"
        elif current_site == "diego" and "CUM" in eu:
            # Diego Cummins = ENELEC + LFO (user spec)
            provider = cfg.PROVIDER_ENELEC
            fuel = "LFO"
        elif current_site == "tamatave" and ("DEUTZ" in make.upper() or eu == "ADG 14"):
            provider = cfg.PROVIDER_VESTOP
        elif current_site == "majunga" and engine.lower() == "vestop":
            provider = cfg.PROVIDER_VESTOP

        statut_low = statut.lower()
        is_running = statut_low == "working"

        out[current_site].append({
            "id":           engine,
            "make":         make,
            "model":        model,
            "nominal":      nominal,
            "attendu":      attendu,
            "availableMw":  avail_num,
            "availableRaw": _s(avail_raw),
            "statut":       statut,
            "statutNorm":   "ok" if is_running else "ko",
            "lubes":        lubes,
            "stopDate":     _iso(stop_date_raw) if not isinstance(stop_date_raw, str) else _s(stop_date_raw),
            "issues":       issues,
            "actions":      actions,
            "dlRevu":       dl_revu,
            "provider":     provider,
            "fuel":         fuel,
        })

    return out


# ══════════════════════════════════════════════════════════════
# HEBDO sheet (weekly planning with colors)
# ══════════════════════════════════════════════════════════════

HEBDO_ENGINE_SECTION_RE = re.compile(
    r"^(ADG|BDG|DDG|MDG|UDG|FDG|DEUTZ|MAN|RR|Cummins)", re.I
)

HEBDO_SITE_HEADERS = {
    "TAMATAVE": "tamatave",
    "MAJUNGA":  "majunga",
    "DIEGO":    "diego",
    "TULEAR":   "tulear",
    "ANTSIRABE": "antsirabe",
    "FIHAONANA": "fihaonana",
}


def load_hebdo_planning(xlsx_path):
    """
    Returns a per-site dict with engine weekly statuses:
        {
          tamatave: {
            weeks: ['2026-01-S1', '2026-01-S2', ...],   # 52 labels
            engines: [
              { id, nominal, provider, weeks: [None/'maintenance'/'indisponible'/'run', ...] }
            ],
            totals: {
              enelec:       [...],  # MW per week or None
              vestop:       [...],
              enelecVestop: [...],
              peakLoad:     [...],
              contrat:      [...],
            }
          },
          ...
        }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[cfg.SHEET_HEBDO]

    # Build the list of week columns by scanning rows 8 (months) + 9 (Sx labels)
    week_cols = []  # [(col_index, iso_week_label)]
    current_month = None
    for c in range(1, ws.max_column + 1):
        m_val = ws.cell(8, c).value
        if isinstance(m_val, datetime):
            current_month = m_val.strftime("%Y-%m")
        s_val = ws.cell(9, c).value
        if isinstance(s_val, str) and s_val.startswith("S") and current_month:
            week_cols.append((c, f"{current_month}-{s_val}"))

    out = {}
    current_site = None
    current_provider = None  # 'enelec' | 'vestop' | 'cummins'

    for r in range(7, ws.max_row + 1):
        label_b = _s(ws.cell(r, 3).value)
        label_c = _s(ws.cell(r, 4).value)

        # Site header
        if label_b.upper() in HEBDO_SITE_HEADERS:
            current_site = HEBDO_SITE_HEADERS[label_b.upper()]
            out.setdefault(current_site, {
                "weeks": [w for _, w in week_cols],
                "engines": [],
                "totals": {
                    "enelec": [None] * len(week_cols),
                    "vestop": [None] * len(week_cols),
                    "enelecVestop": [None] * len(week_cols),
                    "peakLoad": [None] * len(week_cols),
                    "contrat": [None] * len(week_cols),
                },
            })
            current_provider = None
            continue

        if not current_site:
            continue
        site_entry = out[current_site]

        # Provider / section markers
        lb_low = label_b.lower()
        if lb_low in ("enelec", "vestop"):
            # If col C has "Puissance nominale" it's a section header (start of block)
            if label_c.lower().startswith("puissance"):
                current_provider = lb_low
                continue
            # Otherwise it's a total line for that provider
            vals = []
            for (c, _wk) in week_cols:
                vals.append(_f(ws.cell(r, c).value))
            # Stash under the right totals bucket (based on label content)
            if "vestop" in lb_low:
                site_entry["totals"]["vestop"] = vals
            else:
                site_entry["totals"]["enelec"] = vals
            continue

        if "enelec + vestop" in lb_low or "enelec tulear" in lb_low or "enelec diego" in lb_low or "vestop antsirabe" in lb_low:
            site_entry["totals"]["enelecVestop"] = [_f(ws.cell(r, c).value) for (c, _wk) in week_cols]
            continue
        if lb_low == "peak load":
            site_entry["totals"]["peakLoad"] = [_f(ws.cell(r, c).value) for (c, _wk) in week_cols]
            continue
        if lb_low == "contrat":
            site_entry["totals"]["contrat"] = [_f(ws.cell(r, c).value) for (c, _wk) in week_cols]
            continue

        # Cummins/LFO marker
        if lb_low == "cummins":
            current_provider = "cummins"
            continue
        if lb_low == "total load available" or lb_low == "total cummins":
            continue

        # Engine row
        if HEBDO_ENGINE_SECTION_RE.match(label_b):
            nominal = _f(label_c)  # col D = "Puissance nominale" (may be "Cannibalisé" text)
            provider_key = current_provider or "enelec"
            if provider_key == "cummins":
                # LFO on Antsirabe, ENELEC on Diego
                provider_key = cfg.PROVIDER_LFO if current_site == "antsirabe" else cfg.PROVIDER_ENELEC
            weeks_status = []
            for (c, _wk) in week_cols:
                cell = ws.cell(r, c)
                fg = cell.fill.fgColor
                color = fg.rgb if fg and fg.type == "rgb" else None
                status = cfg.PLANNING_STATUS_BY_COLOR.get(color)
                if status is None:
                    # A numeric value means the engine is planned to run that week
                    val = _f(cell.value)
                    status = "run" if val is not None else None
                weeks_status.append(status)
            site_entry["engines"].append({
                "id":       label_b,
                "nominal":  nominal if nominal is not None else label_c,  # keep text if e.g. "Cannibalisé"
                "provider": provider_key,
                "weeks":    weeks_status,
            })

    return out


# ══════════════════════════════════════════════════════════════
# Top-level orchestrator
# ══════════════════════════════════════════════════════════════

def load_global(xlsx_path=None):
    """Load everything from Daily_Report_Global.xlsx and return a dict."""
    path = xlsx_path or cfg.GLOBAL_FILE
    if not os.path.exists(path):
        raise FileNotFoundError(f"Daily_Report_Global.xlsx not found: {path}")

    return {
        "overhaul":   load_overhaul(path),
        "recap":      load_recap(path),
        "situation":  load_situation_moteurs(path),
        "hebdo":      load_hebdo_planning(path),
    }


if __name__ == "__main__":
    # Smoke test
    data = load_global()
    print(f"Overhaul: {len(data['overhaul'])} rows")
    print(f"Recap sites: {list(data['recap'].keys())}")
    print(f"Situation sites: {list(data['situation'].keys())}")
    for site, engines in data["situation"].items():
        print(f"  {site}: {len(engines)} engines")
    print(f"Hebdo sites: {list(data['hebdo'].keys())}")
    for site, entry in data["hebdo"].items():
        print(f"  {site}: {len(entry['weeks'])} weeks, {len(entry['engines'])} engines")
