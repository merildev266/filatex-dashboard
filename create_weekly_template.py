"""
Create Weekly_EnR_Avancement.xlsx template for weekly reporting.
Based on user's Weekly Enr.xlsx structure, adapted for dashboard pipeline.
52 weekly sheets (S01-S52) + Config + PAIEMENT + PAIEMENT STATUS
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ── Source file (user's template with real data) ──
SRC = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\1. DATA Meril HIVANAKO\Bureau\Weekly Enr.xlsx"
OUT_DIR = os.path.join(
    os.environ.get("USERPROFILE", r"C:\Users\Meril"),
    "OneDrive - GROUPE FILATEX",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "01_Energy", "Projet", "EnR", "Weekly_Report",
)
OUT = os.path.join(OUT_DIR, "Weekly_EnR_Avancement.xlsx")

# ── Projects ──
PROJECTS = [
    ("nosy-be-1",      "Nosy Be Phase 1",              "Lindon",       "Solaire",  3),
    ("tulear-2",       "Tulear Phase 2",                "Toki",         "Solaire",  1),
    ("moramanga-1",    "Mandrosolar Phase 1 (MSM)",     "Lindon",       "Solaire",  15),
    ("diego-wind-1",   "Diego Wind Phase 1",            "Toki",         "Eolien",   0.12),
    ("bongatsara-1",   "Top Energie Bongatsara Ph1",    "Henintsoa",    "Solaire",  5),
    ("diego-lidera",   "Lidera Phase 2 Diego",          "Aymar/Kevin",  "Solaire",  7.6),
    ("mahajanga",      "Lidera Phase 2 Majunga",        "Aymar/Kevin",  "Solaire",  10.75),
    ("tamatave",       "Lidera Phase 2 Tamatave",       "Aymar/Kevin",  "Solaire",  16),
    ("oursun-1",      "OurSun ZFI Phase 1",            "Jordy",        "Solaire",  3.2),
    ("oursun-2",      "OurSun ZFI Phase 2",            "Jordy",        "Solaire",  7.8),
    ("bongatsara-2",  "Top Energie Bongatsara Ph2",    "Henintsoa",    "Solaire",  5),
    ("fihaonana-1",   "Vestop Fihaonana Phase 1",      "Toki",         "Solaire",  4),
    ("fihaonana-2",   "Vestop Fihaonana Phase 2",      "Toki",         "Solaire",  15),
    ("diego-wind-2",  "Diego Wind Phase 2",             "Toki",         "Eolien",   4.88),
    ("moramanga-2",   "Mandrosolar Phase 2 (MSM)",      "Lindon",       "Solaire",  25),
    ("nosy-be-2",     "Nosy Be Phase 2",                "Lindon",       "Solaire",  2),
    ("tulear-3",      "Tulear Phase 3",                 "Toki",         "Solaire",  3.1),
    ("marais-masay",  "Top Energie Marais Masay",       "Aymar/Kevin",  "Floating", 10),
    ("ambohidratrimo","Top Energie Ambohidratrimo",     "Toki",         "Solaire",  10),
    ("small-sites",   "Small Sites (4 sites)",          "Aymar/Kevin",  "Solaire",  2.18),
    ("lidera-toamasina-3","Lidera Toamasina Phase 3",   "Aymar",        "Solaire",  2),
]

# ── Styles ──
thin = Side(style='thin')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
title_fill = PatternFill('solid', fgColor='2F5496')
hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='808080')
meta_font = Font(name='Calibri', bold=True, size=11)
meta_fill = PatternFill('solid', fgColor='D9E2F3')
data_font = Font(name='Calibri', size=11)
data_font_bold = Font(name='Calibri', size=11, bold=True)
fill_white = PatternFill('solid', fgColor='FFFFFF')
fill_alt = PatternFill('solid', fgColor='F2F2F2')
fill_green = PatternFill('solid', fgColor='C6EFCE')
font_green = Font(name='Calibri', size=11, color='006100')
font_red = Font(name='Calibri', size=11, color='FF0000')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── Pipeline-compatible headers (row 9) ──
# Cols A-E = static info, F-N = to fill each week
HEADERS = {
    'A': 'ID Projet',
    'B': 'Projet',
    'C': 'Resp.',
    'D': 'Type',
    'E': 'MWc',
    'F': 'Phase',
    'G': 'Avancement (%)',
    'H': 'Glissement (j)',
    'I': 'Statut Etudes',
    'J': 'Statut Construction',
    'K': 'EPC/Contractant',
    'L': 'COD Prevue',
    'M': 'Blocages',
    'N': 'Actions S-1',
    'O': 'Actions Semaine',
    'P': 'Commentaires DG',
}
COL_WIDTHS = {'A':16, 'B':30, 'C':14, 'D':10, 'E':8, 'F':14, 'G':12, 'H':11,
              'I':28, 'J':32, 'K':18, 'L':14, 'M':32, 'N':32, 'O':32, 'P':28}


def read_source_data():
    """Read real data from user's Weekly Enr.xlsx for S11 migration."""
    if not os.path.exists(SRC):
        print(f"  Source not found: {SRC}")
        return {}, [], []
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # Sheet 1: Weekly Reporting Projets — extract status per project
    ws = wb["Weekly Reporting Projets"]
    proj_status = {}
    current_proj = None
    for row in range(10, ws.max_row + 1):
        a = ws.cell(row, 1).value
        if a and str(a).strip():
            current_proj = str(a).strip()
            proj_status[current_proj] = {
                'resp': str(ws.cell(row, 2).value or '').strip(),
                'mwc': str(ws.cell(row, 3).value or '').strip(),
                'status': str(ws.cell(row, 4).value or '').strip(),
                'actions': [],
            }
        if current_proj:
            e = ws.cell(row, 5).value
            if e and str(e).strip():
                proj_status[current_proj]['actions'].append(str(e).strip())

    # Sheet 2: PAIEMENT
    pmt_rows = []
    ws2 = wb["PAIEMENT"]
    for row in range(3, ws2.max_row + 1):
        vals = [ws2.cell(row, c).value for c in range(2, 11)]
        if any(v is not None for v in vals):
            pmt_rows.append(vals)

    # Sheet 3: PAIEMENT STATUS
    inv_rows = []
    ws3 = wb["PAIEMENT STATUS"]
    for row in range(2, ws3.max_row + 1):
        vals = [ws3.cell(row, c).value for c in range(1, 10)]
        if any(v is not None for v in vals):
            inv_rows.append(vals)

    wb.close()
    return proj_status, pmt_rows, inv_rows


def create_weekly_sheet(wb, sheet_name, monday, sunday, is_current=False, src_data=None):
    ws = wb.create_sheet(sheet_name)

    # ── Meta header (rows 1-8) — same style as user's template ──
    ws.merge_cells('A1:P1')
    c = ws['A1']
    c.value = "Weekly Report - Statut d'Avancement Projets EnR"
    c.font = title_font; c.fill = title_fill; c.alignment = center; c.border = thin_border

    ws['A2'] = 'Compile par :'; ws['A2'].font = meta_font; ws['A2'].fill = meta_fill; ws['A2'].border = thin_border
    ws.merge_cells('B2:C2'); ws['B2'].border = thin_border
    ws['D2'] = 'Date :'; ws['D2'].font = meta_font; ws['D2'].fill = meta_fill; ws['D2'].border = thin_border
    ws.merge_cells('E2:F2')
    ws['E2'] = monday; ws['E2'].number_format = 'DD/MM/YYYY'; ws['E2'].font = meta_font; ws['E2'].border = thin_border

    ws['A3'] = 'Semaine :'; ws['A3'].font = meta_font; ws['A3'].fill = meta_fill; ws['A3'].border = thin_border
    ws.merge_cells('B3:C3')
    ws['B3'] = sheet_name; ws['B3'].font = meta_font; ws['B3'].border = thin_border
    ws['D3'] = 'Du :'; ws['D3'].font = meta_font; ws['D3'].fill = meta_fill; ws['D3'].border = thin_border
    ws['E3'] = monday; ws['E3'].number_format = 'DD/MM/YYYY'; ws['E3'].font = meta_font; ws['E3'].border = thin_border
    ws['F3'] = 'Au :'; ws['F3'].font = meta_font; ws['F3'].fill = meta_fill; ws['F3'].border = thin_border
    ws['G3'] = sunday; ws['G3'].number_format = 'DD/MM/YYYY'; ws['G3'].font = meta_font; ws['G3'].border = thin_border

    ws['A4'] = 'Sujet :'; ws['A4'].font = meta_font; ws['A4'].fill = meta_fill; ws['A4'].border = thin_border
    ws.merge_cells('B4:P4')
    ws['B4'] = 'Avancement global projets EnR'; ws['B4'].font = meta_font; ws['B4'].border = thin_border

    ws['A5'] = 'Objectif :'; ws['A5'].font = meta_font; ws['A5'].fill = meta_fill; ws['A5'].border = thin_border
    ws.merge_cells('B5:P5')
    ws['B5'] = "Point sur l'etat des projets - identifier blocages - coordonner actions"; ws['B5'].font = data_font; ws['B5'].border = thin_border

    # Row 6-7 spacer
    ws.merge_cells('A6:P6')
    ws.merge_cells('A7:P7')
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[7].height = 6

    # ── Column headers (row 8) ──
    for col_letter, label in HEADERS.items():
        ci = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(8, ci, label)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border

    # ── Column widths ──
    for col_letter, w in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # ── Project rows (9-29) ──
    for ri, proj in enumerate(PROJECTS):
        r = ri + 9
        fill = fill_white if ri % 2 == 0 else fill_alt
        pid, name, resp, typ, mwc = proj

        row_data = {1: pid, 2: name, 3: resp, 4: typ, 5: mwc}
        for ci in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, ci)
            if ci in row_data:
                cell.value = row_data[ci]
            cell.font = data_font_bold if ci <= 2 else data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center if ci <= 8 else left_wrap

        ws.cell(r, 7).number_format = '0'      # Avancement
        ws.cell(r, 8).number_format = '0'       # Glissement
        ws.cell(r, 12).number_format = 'DD/MM/YYYY'  # COD

    # ── Data validations ──
    dv_type = DataValidation(type='list', formula1='"Solaire,Eolien,Floating"', allow_blank=True)
    dv_phase = DataValidation(type='list', formula1='"Planifie,Developpement,Construction,Termine"', allow_blank=True)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_phase)
    dv_type.add('D9:D29')
    dv_phase.add('F9:F29')

    # ── Freeze & filter ──
    ws.freeze_panes = 'A9'
    ws.auto_filter.ref = f'A8:P29'

    # ── Migrate source data for current week ──
    if is_current and src_data:
        _migrate_data(ws, src_data)


def _migrate_data(ws, src_data):
    """Migrate existing weekly data from old format into current sheet."""
    # Map from old file: read the old Weekly_EnR_Avancement S11
    old_path = os.path.join(OUT_DIR, "Weekly_EnR_Avancement.xlsx")
    # We'll use the src_data (from generate pipeline) which has the old S11 values
    # Instead, read from the existing file if it exists
    try:
        old_wb = openpyxl.load_workbook(old_path, data_only=True)
        if 'S11' in old_wb.sheetnames:
            old_ws = old_wb['S11']
            # Old columns: F=Phase, G=Avancement, H=Glissement, I=StatutEtudes, J=StatutConst,
            # K=PointsCles, L=EPC, M=COD, O=Blocages, P=ActionsS1, Q=ActionsSem
            # New columns: F=Phase, G=Av, H=Glis, I=StatutEtudes, J=StatutConst,
            # K=EPC, L=COD, M=Blocages, N=ActionsS1, O=ActionsSem, P=ComDG
            OLD_NEW = {6:6, 7:7, 8:8, 9:9, 10:10, 12:11, 13:12, 15:13, 16:14, 17:15, 18:16}
            for row in range(5, 26):
                old_pid = old_ws.cell(row, 1).value
                if not old_pid: continue
                # Find matching row in new sheet
                for nr in range(9, 30):
                    if ws.cell(nr, 1).value == str(old_pid).strip():
                        for old_c, new_c in OLD_NEW.items():
                            v = old_ws.cell(row, old_c).value
                            if v is not None:
                                ws.cell(nr, new_c).value = v
                        break
        old_wb.close()
    except Exception as e:
        print(f"  Migration note: {e}")


def create_paiement_sheet(wb, pmt_rows):
    ws = wb.create_sheet("PAIEMENT")
    # Headers
    pmt_hdrs = ['PROJET','CONTRACTANT','PRESTATION','DESCRIPTION','PROPORTION','Montant $','Echeance','Statut','Actions en cours']
    pmt_cols = ['B','C','D','E','F','G','H','I','J']
    pmt_widths = {'A':4, 'B':15, 'C':17, 'D':14, 'E':32, 'F':10, 'G':16, 'H':12, 'I':10, 'J':44}
    orange_fill = PatternFill('solid', fgColor='ED7D31')

    for col, w in pmt_widths.items():
        ws.column_dimensions[col].width = w

    for i, (col, hdr) in enumerate(zip(pmt_cols, pmt_hdrs)):
        ci = openpyxl.utils.column_index_from_string(col)
        cell = ws.cell(2, ci, hdr)
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.fill = orange_fill
        cell.alignment = center
        cell.border = thin_border

    # Data
    for ri, vals in enumerate(pmt_rows):
        r = ri + 3
        for ci_off, v in enumerate(vals):
            ci = ci_off + 2  # starts at col B
            cell = ws.cell(r, ci)
            cell.value = v
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_wrap
            # Statut formatting
            if ci == 9:  # col I
                sv = str(v or '').strip().upper()
                if sv == 'OK':
                    cell.fill = fill_green; cell.font = font_green
                elif sv == 'NOK':
                    cell.font = font_red
        ws.cell(r, 6).number_format = '0%'
        ws.cell(r, 7).number_format = '#,##0.00'
        ws.cell(r, 8).number_format = 'mm-dd-yy'

    ws.freeze_panes = 'A3'


def create_paiement_status_sheet(wb, inv_rows):
    ws = wb.create_sheet("PAIEMENT STATUS")
    inv_hdrs = ['Ref FA','Montant','Currency','Fournisseur','Date de facture','Echeance','Projet','Site','Libelle']
    inv_widths = {'A':25, 'B':11, 'C':10, 'D':14, 'E':15, 'F':11, 'G':25, 'H':11, 'I':58}

    for col, w in inv_widths.items():
        ws.column_dimensions[col].width = w

    for ci, hdr in enumerate(inv_hdrs, 1):
        cell = ws.cell(1, ci, hdr)
        cell.font = data_font
        cell.border = thin_border

    for ri, vals in enumerate(inv_rows):
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri + 2, ci)
            cell.value = v
            cell.font = data_font
        ws.cell(ri + 2, 5).number_format = 'mm-dd-yy'
        ws.cell(ri + 2, 6).number_format = 'mm-dd-yy'


def main():
    print("Creating Weekly_EnR_Avancement template...")
    print(f"  Source: {SRC}")
    print(f"  Output: {OUT}")

    # Read source data
    proj_status, pmt_rows, inv_rows = read_source_data()
    print(f"  Source projects: {len(proj_status)}, payments: {len(pmt_rows)}, invoices: {len(inv_rows)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 52 weekly sheets
    year_start = datetime(2025, 12, 29)  # Monday of week 1
    current_week = 12  # S12 = week of 16/03/2026

    for wk in range(1, 53):
        monday = year_start + timedelta(weeks=wk - 1)
        sunday = monday + timedelta(days=6)
        sn = f'S{wk:02d}'
        is_current = (wk == 11)  # S11 has existing data to migrate
        create_weekly_sheet(wb, sn, monday, sunday, is_current=is_current)
        if wk % 10 == 0:
            print(f"  Created {sn}...")

    # Config sheet
    ws_cfg = wb.create_sheet('Config')
    ws_cfg['A1'] = 'Configuration'; ws_cfg['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws_cfg['A3'] = 'Annee'; ws_cfg['B3'] = 2026
    ws_cfg['A4'] = 'Semaine active'; ws_cfg['B4'] = f'S{current_week:02d}'
    ws_cfg['A5'] = 'Nb Projets'; ws_cfg['B5'] = len(PROJECTS)
    ws_cfg['A7'] = 'ID Projet'; ws_cfg['B7'] = 'Nom'; ws_cfg['C7'] = 'Responsable'
    ws_cfg['A7'].font = Font(bold=True); ws_cfg['B7'].font = Font(bold=True); ws_cfg['C7'].font = Font(bold=True)
    for i, p in enumerate(PROJECTS):
        ws_cfg.cell(8 + i, 1, p[0])
        ws_cfg.cell(8 + i, 2, p[1])
        ws_cfg.cell(8 + i, 3, p[2])
    ws_cfg.column_dimensions['A'].width = 22
    ws_cfg.column_dimensions['B'].width = 35
    ws_cfg.column_dimensions['C'].width = 15

    # PAIEMENT + PAIEMENT STATUS sheets (from source)
    if pmt_rows:
        create_paiement_sheet(wb, pmt_rows)
        print(f"  PAIEMENT: {len(pmt_rows)} rows")
    if inv_rows:
        create_paiement_status_sheet(wb, inv_rows)
        print(f"  PAIEMENT STATUS: {len(inv_rows)} rows")

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT)
    print(f"\nOK: {OUT}")
    print(f"  Sheets: {len(wb.sheetnames)} (52 weeks + Config + PAIEMENT + PAIEMENT STATUS)")


if __name__ == "__main__":
    main()
