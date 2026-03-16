"""Generate reporting_data.js from Weekly EnR Excel files (52-sheet format).
Exports ALL weeks with data for client-side week switching."""
import json, os, sys, re
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

BASE = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\01_Energy\Projet\EnR"
AVANCEMENT_FILE = os.path.join(BASE, "Weekly_Report", "Weekly_EnR_Avancement.xlsx")
PAIEMENTS_FILE = os.path.join(BASE, "Weekly_Report", "Weekly_EnR_Paiements.xlsx")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reporting_data.js")


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def safe_num(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _find_latest_sheet(wb):
    """Find the latest filled S## sheet in a workbook."""
    if "Config" in wb.sheetnames:
        cfg_val = wb["Config"]["B4"].value
        if cfg_val and re.match(r"^S\d{2}$", str(cfg_val)):
            return str(cfg_val)

    latest = None
    for sn in wb.sheetnames:
        m = re.match(r"^S(\d{2})$", sn)
        if not m:
            continue
        ws = wb[sn]
        has_data = False
        for r in range(9, min(ws.max_row + 1, 35)):
            for c in range(6, min(ws.max_column + 1, 18)):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    has_data = True
                    break
            if has_data:
                break
        if has_data:
            num = int(m.group(1))
            if latest is None or num > latest:
                latest = num
    return f"S{latest:02d}" if latest else None


def _find_all_filled_sheets(wb):
    """Find all S## sheets that have real data.
    Checks cols 6-15 (Phase through Actions) — excludes comment-only sheets."""
    filled = []
    for sn in wb.sheetnames:
        m = re.match(r"^S(\d{2})$", sn)
        if not m:
            continue
        ws = wb[sn]
        has_data = False
        for r in range(9, min(ws.max_row + 1, 35)):
            for c in range(6, 16):  # F=Phase through O=Actions (not P=ComDG)
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    has_data = True
                    break
            if has_data:
                break
        if has_data:
            filled.append(sn)
    return filled


def _extract_week_date(ws):
    """Extract week date string from sheet header cells."""
    for cell_ref in ["G3", "E3", "E2"]:
        val = ws[cell_ref].value
        if val is None:
            continue
        if hasattr(val, 'strftime'):
            return val.strftime("%d/%m/%Y")
        s = safe_str(val)
        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", s)
        if date_match:
            return date_match.group(1)
    return ""


def _read_sheet_projects(ws):
    """Read project rows from a weekly sheet."""
    projects = []
    for row in range(9, ws.max_row + 1):
        pid = safe_str(ws.cell(row, 1).value)
        nom = safe_str(ws.cell(row, 2).value)
        if not pid or not nom:
            continue
        projects.append({
            "id": pid,
            "projet": nom,
            "responsable": safe_str(ws.cell(row, 3).value),
            "type": safe_str(ws.cell(row, 4).value),
            "puissance": safe_num(ws.cell(row, 5).value),
            "phase": safe_str(ws.cell(row, 6).value),
            "avancement": safe_num(ws.cell(row, 7).value),
            "glissement": int(safe_num(ws.cell(row, 8).value)),
            "statut_eng": safe_str(ws.cell(row, 9).value),
            "statut_const": safe_str(ws.cell(row, 10).value),
            "statut_permis": "",
            "epc": safe_str(ws.cell(row, 11).value),
            "date_mes_prevue": safe_str(ws.cell(row, 12).value),
            "date_mes_revisee": "",
            "blocages": safe_str(ws.cell(row, 13).value),
            "actions_s1": safe_str(ws.cell(row, 14).value),
            "actions": safe_str(ws.cell(row, 15).value),
            "commentaires_dg": safe_str(ws.cell(row, 16).value),
            "reponse": safe_str(ws.cell(row, 17).value),
        })
    return projects


def read_all_weeks():
    """Read ALL filled weeks from the avancement file."""
    if not os.path.exists(AVANCEMENT_FILE):
        print(f"WARN: {AVANCEMENT_FILE} not found")
        return None, {}, []

    wb = openpyxl.load_workbook(AVANCEMENT_FILE, data_only=True)

    # Check if it's 52-sheet format
    filled_sheets = _find_all_filled_sheets(wb)
    if filled_sheets:
        weeks_data = {}
        latest_sheet = None
        latest_num = 0
        # Check Config for explicit latest
        if "Config" in wb.sheetnames:
            cfg_val = wb["Config"]["B4"].value
            if cfg_val and re.match(r"^S\d{2}$", str(cfg_val)):
                latest_sheet = str(cfg_val)
                latest_num = int(str(cfg_val)[1:])

        for sn in filled_sheets:
            ws = wb[sn]
            week_date = _extract_week_date(ws)
            projects = _read_sheet_projects(ws)
            num = int(sn[1:])
            weeks_data[sn] = {
                "sheet": sn,
                "week": week_date,
                "projects": projects,
            }
            if not latest_sheet and num > latest_num:
                latest_num = num
                latest_sheet = sn

        if not latest_sheet and filled_sheets:
            latest_sheet = filled_sheets[-1]

        wb.close()
        return latest_sheet, weeks_data, []

    # Fallback to old single-sheet format
    if "Avancement Projets" in wb.sheetnames:
        ws = wb["Avancement Projets"]
        week_val = ws["E2"].value
        week_str = ""
        if week_val:
            week_str = week_val.strftime("%d/%m/%Y") if hasattr(week_val, 'strftime') else str(week_val)
        projects = []
        for row in range(5, ws.max_row + 1):
            pid = safe_str(ws.cell(row, 1).value)
            nom = safe_str(ws.cell(row, 2).value)
            if not pid and not nom:
                continue
            projects.append({
                "id": pid, "projet": nom,
                "responsable": safe_str(ws.cell(row, 3).value),
                "type": safe_str(ws.cell(row, 4).value),
                "puissance": safe_num(ws.cell(row, 5).value),
                "phase": safe_str(ws.cell(row, 6).value),
                "avancement": safe_num(ws.cell(row, 7).value),
                "glissement": int(safe_num(ws.cell(row, 8).value)),
                "statut_eng": safe_str(ws.cell(row, 9).value),
                "statut_const": safe_str(ws.cell(row, 10).value),
                "statut_permis": safe_str(ws.cell(row, 11).value),
                "epc": safe_str(ws.cell(row, 12).value),
                "date_mes_prevue": safe_str(ws.cell(row, 13).value),
                "date_mes_revisee": safe_str(ws.cell(row, 14).value),
                "blocages": safe_str(ws.cell(row, 15).value),
                "actions_s1": safe_str(ws.cell(row, 16).value),
                "actions": safe_str(ws.cell(row, 17).value),
                "commentaires_dg": safe_str(ws.cell(row, 18).value),
                "reponse": safe_str(ws.cell(row, 19).value),
            })
        wb.close()
        weeks_data = {"S00": {"sheet": "S00", "week": week_str, "projects": projects}}
        return "S00", weeks_data, []

    wb.close()
    return None, {}, []


def read_paiements():
    if not os.path.exists(PAIEMENTS_FILE):
        print(f"WARN: {PAIEMENTS_FILE} not found")
        return []

    wb = openpyxl.load_workbook(PAIEMENTS_FILE, data_only=True)

    sheet_name = _find_latest_sheet(wb)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row = 5
    elif "Paiements" in wb.sheetnames:
        ws = wb["Paiements"]
        start_row = 5
    else:
        wb.close()
        print("WARN: No suitable paiements sheet found")
        return []

    payments = []
    for row in range(start_row, ws.max_row + 1):
        pid = safe_str(ws.cell(row, 1).value)
        projet = safe_str(ws.cell(row, 2).value)
        if not pid and not projet:
            continue
        if safe_str(ws.cell(row, 5).value).startswith("TOTAL"):
            continue
        payments.append({
            "id_projet": pid, "projet": projet,
            "contractant": safe_str(ws.cell(row, 3).value),
            "prestation": safe_str(ws.cell(row, 4).value),
            "description": safe_str(ws.cell(row, 5).value),
            "montant": safe_num(ws.cell(row, 6).value),
            "devise": safe_str(ws.cell(row, 7).value),
            "echeance": safe_str(ws.cell(row, 8).value),
            "statut": safe_str(ws.cell(row, 9).value),
            "actions": safe_str(ws.cell(row, 10).value),
        })

    wb.close()
    return payments


def main():
    latest_sheet, weeks_data, _ = read_all_weeks()
    payments = read_paiements()

    # Build backward-compatible format + weeks index
    latest_data = weeks_data.get(latest_sheet, {})

    data = {
        "week": latest_data.get("week", ""),
        "currentSheet": latest_sheet or "",
        "generated": __import__('datetime').datetime.now().isoformat(),
        "projects": latest_data.get("projects", []),
        "payments": payments,
        "weeks": weeks_data,
    }

    js = "/* Auto-generated by generate_reporting.py — DO NOT EDIT */\nwindow.REPORTING_ENR = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"

    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"OK: {OUT}")
    print(f"  Latest: {latest_sheet}")
    print(f"  Weeks with data: {len(weeks_data)} ({', '.join(sorted(weeks_data.keys()))})")
    print(f"  Projects (latest): {len(latest_data.get('projects', []))}")
    print(f"  Payments: {len(payments)}")


if __name__ == "__main__":
    main()
