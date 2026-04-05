"""Migrate existing Investments data from old format (1 sheet/project) into new weekly template."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OLD = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\03_ Investments\Reporting\Avancement (1).xlsx"
TPL = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\03_ Investments\Reporting\Weekly_Investments_Avancement_v2.xlsx"

# Map old sheet names to new project IDs
SHEET_TO_PID = {
    "Oasis":              "oasis",
    "Orga Earth":         "orga-earth",
    "Hakanto House":      "hakanto-house",
    "MLF":                "mlf",
    "Energiestro":        "energiestro",
    "Hotel Tamatave":     "hotel-tamatave",
    "SunFarming":         "sunfarming",
    "Afridoctor":         "afridoctor",
    "Artemis":            "artemis",
    "BGFI":               "bgfi",
    "Oui Coding":         "oui-coding",
    "Seed Star":          "seed-star",
    "Café Mary":          "cafe-mary",
    "Caf\u00e9 Mary":     "cafe-mary",
    "GHU ":               "ghu",
    "GHU":                "ghu",
    "Haya":               "haya",
    "Maison des cotonniers": "maison-cotonniers",
    "Show Room":          "show-room",
    "SSLS":               "ssls",
    "Taxi Brousse Pizza": "taxi-brousse-pizza",
}

# Load old workbook (data_only to get calculated values)
old_wb = openpyxl.load_workbook(OLD, data_only=True)
# Load new template
new_wb = openpyxl.load_workbook(TPL)

# For each old sheet, extract weekly data and inject into new template
total_filled = 0

for old_sheet_name in old_wb.sheetnames:
    old_ws = old_wb[old_sheet_name]
    pid = SHEET_TO_PID.get(old_sheet_name)
    if not pid:
        # Try stripping
        pid = SHEET_TO_PID.get(old_sheet_name.strip())
    if not pid:
        print(f"SKIP: No mapping for sheet '{old_sheet_name}'")
        continue

    # Read weekly rows from old sheet (rows 7-58, col A=week num)
    for r in range(7, 59):
        wk = old_ws.cell(r, 1).value
        if not wk or not isinstance(wk, (int, float)):
            continue
        wk = int(wk)
        statut_val = old_ws.cell(r, 3).value
        av_val = old_ws.cell(r, 4).value
        bloc_val = old_ws.cell(r, 5).value
        maj_val = old_ws.cell(r, 6).value

        # Skip empty weeks
        if not any([statut_val, av_val, bloc_val]):
            continue

        # Find target sheet
        sname = f"S{wk:02d}"
        if sname not in new_wb.sheetnames:
            continue
        new_ws = new_wb[sname]

        # Find the row with matching project ID in new sheet
        target_row = None
        for nr in range(6, 35):
            cell_val = new_ws.cell(nr, 1).value
            if cell_val == pid:
                target_row = nr
                break

        if not target_row:
            continue

        # New layout: D=Statut, E=Avancement, F=Blocage, I=Mise a jour
        if statut_val:
            sv = str(statut_val).strip()
            if sv:
                new_ws.cell(target_row, 4).value = sv
        if av_val:
            new_ws.cell(target_row, 5).value = str(av_val).strip()
        if bloc_val:
            new_ws.cell(target_row, 6).value = str(bloc_val).strip()
        if maj_val:
            new_ws.cell(target_row, 9).value = str(maj_val).strip()

        total_filled += 1

# Update config
if "Config" in new_wb.sheetnames:
    new_wb["Config"]["B4"] = "S11"

new_wb.save(TPL)
print(f"Migration done: {total_filled} cells filled across all weeks")
print("Saved OK")
