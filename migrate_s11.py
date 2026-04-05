"""Migrate S11 data into the new Weekly template."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

TPL = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\01_Energy\Projet\EnR\Weekly_Report\Weekly_EnR_Avancement.xlsx"

OLD_DATA = {
    "nosy-be-1":    {"phase":"Termine",       "av":100, "glis":168, "epc":"SQVISION",     "etudes":"100%", "const":"100% - MES 07/02/2026", "bloc":"Levee de reserves en attente", "actS1":"MES effectuee le 07/02/2026", "actS":"Paiement EPC solde"},
    "tulear-2":     {"phase":"Termine",       "av":98,  "glis":61,  "epc":"MADAGREEN",    "etudes":"100%", "const":"98% - COD en cours", "bloc":"Reserves non levees - retour Socotec", "actS1":"Retour Socotec recu", "actS":"Action Madagreen planning levee reserves"},
    "moramanga-1":  {"phase":"Construction",  "av":67,  "glis":135, "epc":"BLUESKY",      "etudes":"100%", "const":"67% - Pose modules 6.7MW", "bloc":"Retard dedouanement + logistique route", "actS1":"Pose module solaire 6.7MW", "actS":"Poursuite installation structures + modules"},
    "diego-wind-1": {"phase":"Construction",  "av":61,  "glis":208, "epc":"AEOLOS/OTI",   "etudes":"90%",  "const":"61% - Installation en cours", "bloc":"Permis MEH bloque + retard paiement dedouanement", "actS1":"OTI 3e/4e paiement OV signe", "actS":"Dedouanement 10140$ + magasinage 14840$; Relance appel offres 3e shipment"},
    "bongatsara-1": {"phase":"Construction",  "av":26,  "glis":0,   "epc":"CERMAD/TBC",   "etudes":"80%",  "const":"GC 47% - ARRET TRAVAUX 28/02", "bloc":"ARRET TRAVAUX CERMAD - retard paiement", "actS1":"Analyse technique EPC Grid Tech vs Bluesky", "actS":"Reprise paiement CERMAD; Validation SOCOTEC; Choix EPC installateur"},
    "diego-lidera": {"phase":"Developpement", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"Etudes Geotech validees", "bloc":"Redimensionnement puissances besoins reels", "actS1":"Etude Geotech LAGEOTECH 100%", "actS":"Etude offres leasing Ithemba + Scatec; PPA execution"},
    "mahajanga":    {"phase":"Developpement", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"Etudes Geotech validees", "bloc":"Redimensionnement puissances besoins reels", "actS1":"Etude Geotech BRG 100%", "actS":"Paiement solde 50% BRG; Etude offres leasing Ithemba"},
    "tamatave":     {"phase":"Developpement", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"Etudes Geotech en cours", "bloc":"Discussion terrain en cours (Immo)", "actS1":"Contrat etude geotech signe", "actS":"Relance Immo terrain; Paiement stand-by attente foncier"},
    "oursun-1":     {"phase":"Developpement", "av":10,  "glis":0,   "epc":"BLUESKY",      "etudes":"70% - Livrable provisoire recu", "bloc":"Loyer IMMO impaye 51164 USD (Jan-Dec 2025)", "actS1":"Etude faisabilite Artelia 70%", "actS":"Debut travaux 09/06/2026; HSE preparation; Structuration partenariat BlueSky"},
    "oursun-2":     {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "bloc":"Renforcement toiture en cours", "actS":"Attente toiture renforcee"},
    "bongatsara-2": {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "bloc":"Changement terrain vers Ambohijanaka", "actS":"Terrain pas encore defini"},
    "fihaonana-1":  {"phase":"Developpement", "av":24,  "glis":0,   "epc":"TBC",          "etudes":"Faisabilite Artelia 90%", "bloc":"Attente offre ligne evacuation JIRAMA", "actS1":"Consultation EPC lancee DL 27/03", "actS":"Depot permis env.; Paiement Geo-Eco 70% restant; Validation controle SOCOTEC"},
    "fihaonana-2":  {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "actS":"Attente offres etude faisabilite + geotech; DL validation env. 18/03"},
    "diego-wind-2": {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "etudes":"Consultation lancee", "actS":"Consultation etude faisabilite; Etude Topo debut avril"},
    "moramanga-2":  {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "etudes":"35%", "actS":"Debut travaux 30/06/2026; CDC a mettre a jour pour consultation EPC"},
    "nosy-be-2":    {"phase":"Developpement", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"90% - Etudes completees", "bloc":"Paiement permis en attente signature contrat EPC", "actS":"Offres EPC PV+BESS recues; Debut travaux 01/05/26"},
    "tulear-3":     {"phase":"Developpement", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"100% - Etude effectuee", "actS1":"Comparatif offres EPC en cours", "actS":"Appel offres EPC (7 soumissionnaires); Validation SOCOTEC"},
    "marais-masay": {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"TBC",          "etudes":"En cours", "bloc":"Devis BRG en attente validation DG depuis 27/01", "actS1":"Envoi docs techniques NTPC par Sayouba", "actS":"Acquisition lac - procedure fonciere; Confirmation JIRAMA ligne evacuation"},
    "ambohidratrimo":{"phase":"Planifie",     "av":0,   "glis":0,   "epc":"TBC",          "bloc":"Changement terrain", "actS":"Validation terrain Ambohimalaza pour 5MW"},
    "small-sites":  {"phase":"Planifie",      "av":0,   "glis":0,   "epc":"OTC/BLUESKY",  "etudes":"Etudes preliminaires", "bloc":"Terrains indisponibles Ihosy/Ranohira", "actS":"Offres leasing recues 3 sites; Consultation BlueSky envoyee 27/02; Confirmation descente sites"},
    "lidera-toamasina-3":{"phase":"Planifie", "av":0,   "glis":0,   "epc":"TBC",          "etudes":"0%"},
}

wb = openpyxl.load_workbook(TPL)
ws = wb["S11"]

# Col mapping: F=Phase G=Av H=Glis I=Etudes J=Const K=EPC L=COD M=Bloc N=ActS1 O=ActS P=ComDG
filled = 0
for row in range(9, 30):
    pid = ws.cell(row, 1).value
    if not pid or pid not in OLD_DATA:
        continue
    d = OLD_DATA[pid]
    ws.cell(row, 6).value = d.get("phase")
    ws.cell(row, 7).value = d.get("av")
    ws.cell(row, 8).value = d.get("glis")
    ws.cell(row, 9).value = d.get("etudes", "")
    ws.cell(row, 10).value = d.get("const", "")
    ws.cell(row, 11).value = d.get("epc", "")
    ws.cell(row, 13).value = d.get("bloc", "")
    ws.cell(row, 14).value = d.get("actS1", "")
    ws.cell(row, 15).value = d.get("actS", "")
    filled += 1

if "Config" in wb.sheetnames:
    wb["Config"]["B4"] = "S11"

wb.save(TPL)
print(f"Filled S11: {filled} projects")
print("Saved OK")
