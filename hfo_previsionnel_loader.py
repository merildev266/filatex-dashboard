"""
Parses Previsionnel_26.xlsx.

This file stores the *forecast* production, SFOC and SLOC for every HFO/VESTOP
site, one column per day, May 2025 → end 2026 (~697 cols).

Every sheet has a slightly different row layout, so we keep a per-site map
instead of assuming a global one.

Common conventions:
  - Row 5 (1-indexed) holds the day-of-month number in each day column.
  - Between week blocks there is a NaN separator column (skipped naturally).
  - Data starts at May 6, 2025 (first column with a day number).
  - Day-number resets (day N → day 1) mark month boundaries.

Output per site:
    {
      'monthly': {'2026-01': {prodEnelec, prodVestop, prodLfo, prodSolar,
                              prodTotal, sfocAvg, slocAvg, days}, ...},
      'year2025': {...},
      'year2026': {...},
    }

TANA sheet is ignored per user request.
"""
import os

from openpyxl import load_workbook

import hfo_config as cfg


# Day number row and first data column (same in every PP sheet)
ROW_DAY        = 5
FIRST_DATA_COL = 4

# Anchor date (file starts at Mai 2025 week 2 — May 6)
ANCHOR_YEAR  = 2025
ANCHOR_MONTH = 5


# Per-site row layout (1-indexed). None = not applicable for this site.
SITE_ROWS = {
    "tamatave": {
        "enelec": 19, "vestop": 21, "lfo": None, "solar": 22, "total": 23,
        "sloc_site": 39, "sfoc_site": 53,
        "sloc_vestop": 56, "sfoc_vestop": 57,
    },
    "majunga": {
        "enelec": 11, "vestop": 15, "lfo": 19, "solar": 21, "total": 22,
        "sloc_site": 30, "sfoc_site": 36,
        "sloc_vestop": 40, "sfoc_vestop": 43,
    },
    "diego": {
        "enelec": 15, "vestop": None, "lfo": 23, "solar": 25, "total": 26,
        "sloc_site": 36, "sfoc_site": 45,
        "sloc_vestop": None, "sfoc_vestop": None,
    },
    "tulear": {
        "enelec": 10, "vestop": None, "lfo": None, "solar": 12, "total": 13,
        "sloc_site": 19, "sfoc_site": 24,
        "sloc_vestop": None, "sfoc_vestop": None,
    },
    "antsirabe": {
        "enelec": None, "vestop": 9, "lfo": 15, "solar": None, "total": 17,
        "sloc_site": 19, "sfoc_site": 20,
        "sloc_vestop": None, "sfoc_vestop": None,
    },
}


def _f(v):
    try:
        if v is None:
            return None
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def _col_date_iter(ws):
    """Walk row 5 and yield (col, year, month, day) for every data column."""
    year, month = ANCHOR_YEAR, ANCHOR_MONTH
    prev_day = None
    for c in range(FIRST_DATA_COL, ws.max_column + 1):
        v = ws.cell(ROW_DAY, c).value
        if not isinstance(v, (int, float)) or isinstance(v, bool):
            continue
        day = int(v)
        if prev_day is not None and day <= prev_day:
            month += 1
            if month > 12:
                month = 1
                year += 1
        yield (c, year, month, day)
        prev_day = day


def _read_row(ws, row, c):
    if row is None:
        return None
    return _f(ws.cell(row, c).value)


def load_site_previsionnel(ws, site_key):
    """Load forecasts for a single PP sheet (site_key selects the row map)."""
    rows = SITE_ROWS.get(site_key)
    if rows is None:
        return None

    monthly = {}
    for c, y, m, _d in _col_date_iter(ws):
        key = f"{y}-{m:02d}"
        slot = monthly.setdefault(key, {
            "prodEnelec": 0.0, "prodVestop": 0.0, "prodLfo": 0.0, "prodSolar": 0.0,
            "prodTotal": 0.0,
            "sfoc_sum": 0.0, "sfoc_count": 0,
            "sloc_sum": 0.0, "sloc_count": 0,
            "days": 0,
        })
        slot["days"] += 1

        for label, row_key in (
            ("prodEnelec", "enelec"),
            ("prodVestop", "vestop"),
            ("prodLfo",    "lfo"),
            ("prodSolar",  "solar"),
            ("prodTotal",  "total"),
        ):
            v = _read_row(ws, rows[row_key], c)
            if v is not None:
                slot[label] += v

        sfoc = _read_row(ws, rows["sfoc_site"], c)
        sloc = _read_row(ws, rows["sloc_site"], c)
        if sfoc is not None and sfoc > 0:
            slot["sfoc_sum"] += sfoc
            slot["sfoc_count"] += 1
        if sloc is not None and sloc > 0:
            slot["sloc_sum"] += sloc
            slot["sloc_count"] += 1

    # Serialize
    monthly_out = {}
    for k, slot in monthly.items():
        monthly_out[k] = {
            "prodEnelec": round(slot["prodEnelec"], 2),
            "prodVestop": round(slot["prodVestop"], 2),
            "prodLfo":    round(slot["prodLfo"], 2),
            "prodSolar":  round(slot["prodSolar"], 2),
            "prodTotal":  round(slot["prodTotal"], 2),
            "sfocAvg":    round(slot["sfoc_sum"] / slot["sfoc_count"], 1) if slot["sfoc_count"] else None,
            "slocAvg":    round(slot["sloc_sum"] / slot["sloc_count"], 2) if slot["sloc_count"] else None,
            "days":       slot["days"],
        }

    def _year_totals(year):
        prefix = f"{year}-"
        rows_ = [v for k, v in monthly_out.items() if k.startswith(prefix)]
        if not rows_:
            return None
        sfocs = [r["sfocAvg"] for r in rows_ if r["sfocAvg"] is not None]
        slocs = [r["slocAvg"] for r in rows_ if r["slocAvg"] is not None]
        return {
            "prodEnelec": round(sum(r["prodEnelec"] for r in rows_), 2),
            "prodVestop": round(sum(r["prodVestop"] for r in rows_), 2),
            "prodLfo":    round(sum(r["prodLfo"]    for r in rows_), 2),
            "prodSolar":  round(sum(r["prodSolar"]  for r in rows_), 2),
            "prodTotal":  round(sum(r["prodTotal"]  for r in rows_), 2),
            "sfocAvg":    round(sum(sfocs) / len(sfocs), 1) if sfocs else None,
            "slocAvg":    round(sum(slocs) / len(slocs), 2) if slocs else None,
        }

    return {
        "monthly":  monthly_out,
        "year2025": _year_totals(2025),
        "year2026": _year_totals(2026),
    }


def load_previsionnel(xlsx_path=None):
    """Load the whole Previsionnel_26.xlsx and return per-site dicts."""
    path = xlsx_path or cfg.PREVISIONNEL_FILE
    if not os.path.exists(path):
        raise FileNotFoundError(f"Previsionnel_26.xlsx not found: {path}")

    wb = load_workbook(path, data_only=True)
    out = {}
    for site_key, sheet_name in cfg.PREVISIONNEL_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        data = load_site_previsionnel(ws, site_key)
        if data is not None:
            out[site_key] = data
    return out


if __name__ == "__main__":
    data = load_previsionnel()
    for site, d in data.items():
        y26 = d.get("year2026") or {}
        print(f"{site:10s} 2026: prodTotal={y26.get('prodTotal')} MWh  "
              f"enelec={y26.get('prodEnelec')}  vestop={y26.get('prodVestop')}  lfo={y26.get('prodLfo')}  "
              f"sfoc={y26.get('sfocAvg')}  sloc={y26.get('slocAvg')}")
