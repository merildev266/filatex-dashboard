import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

# Color palette
GREEN = "00AB63"
DARK_BG = "0A0E1A"
HEADER_BG = "1A1F33"
ROW_ALT = "0F1322"
WHITE = "FFFFFF"
DIM = "8A92AB"
ORANGE = "FDB823"
RED = "E05C5C"
CYAN = "5AAFAF"

hdr_font = Font(name="Arial", bold=True, size=9, color=WHITE)
hdr_fill = PatternFill("solid", fgColor=HEADER_BG)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="2A2F44"),
    right=Side(style="thin", color="2A2F44"),
    top=Side(style="thin", color="2A2F44"),
    bottom=Side(style="thin", color="2A2F44"),
)
data_font = Font(name="Arial", size=9, color=WHITE)
data_fill1 = PatternFill("solid", fgColor=DARK_BG)
data_fill2 = PatternFill("solid", fgColor=ROW_ALT)
dim_font = Font(name="Arial", size=8, color=DIM)
green_font = Font(name="Arial", bold=True, size=9, color=GREEN)
title_font = Font(name="Arial", bold=True, size=14, color=GREEN)
subtitle_font = Font(name="Arial", size=10, color=DIM)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.fill = data_fill2 if alt else data_fill1
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def apply_dark_bg(ws):
    for row in range(1, ws.max_row + 3):
        for col in range(1, ws.max_column + 2):
            cell = ws.cell(row=row, column=col)
            if cell.fill == PatternFill() or (cell.fill.fgColor and cell.fill.fgColor.rgb == "00000000"):
                cell.fill = PatternFill("solid", fgColor=DARK_BG)


# Project data shared between both files
projects = [
    ["nosy-be-1", "Nosy-Be Phase 1", "Lindon", "Solaire", 3, "Termine", 100, 168,
     "100%", "100% - MES 07/02/2026", "OK", "SQVISION", "2025-08-23", "2026-02-07",
     "Levee de reserves en attente", "MES effectuee le 07/02/2026", "Paiement EPC solde", ""],
    ["tulear-2", "Tulear Phase 2", "Toki", "Solaire", 1, "Termine", 98, 61,
     "100%", "98% - COD en cours", "OK", "MADAGREEN", "2025-07-23", "2025-09-22",
     "Reserves non levees - retour Socotec", "Retour Socotec recu", "Action Madagreen planning levee reserves", ""],
    ["moramanga-1", "Moramanga Phase 1 (MSM)", "Lindon", "Solaire", 15, "Construction", 67, 135,
     "100%", "67% - Pose modules 6.7MW", "OK", "BLUESKY", "2025-12-01", "2026-05-15",
     "Retard dedouanement + logistique route", "Pose module solaire 6.7MW", "Poursuite installation structures + modules", ""],
    ["diego-wind-1", "Diego Wind Phase 1", "Toki", "Eolien", 0.12, "Construction", 61, 208,
     "90%", "61% - Installation en cours", "Blocage permis MEH", "AEOLOS/OTI", "2025-11-02", "2026-06-25",
     "Permis MEH bloque + retard paiement dedouanement", "OTI 3e/4e paiement OV signe", "Dedouanement 10140$ + magasinage 14840$; Relance appel offres 3e shipment", ""],
    ["bongatsara-1", "Bongatsara Phase 1", "Henintsoa", "Solaire", 5, "Construction", 26, 0,
     "80%", "GC 47% - ARRET TRAVAUX 28/02", "Categorisation recue 09/12", "CERMAD/TBC", "2026-12-30", "",
     "ARRET TRAVAUX CERMAD - retard paiement", "Analyse technique EPC Grid Tech vs Bluesky", "Reprise paiement CERMAD; Validation SOCOTEC; Choix EPC installateur", ""],
    ["diego-lidera", "Lidera Diego Phase 2", "Aymar", "Solaire", 7.6, "Developpement", 0, 0,
     "Etudes Geotech validees", "", "", "TBC", "2027-01-16", "",
     "Redimensionnement puissances besoins reels", "Etude Geotech LAGEOTECH 100%", "Etude offres leasing Ithemba + Scatec; PPA execution", ""],
    ["mahajanga", "Lidera Majunga Phase 2", "Aymar", "Solaire", 10.75, "Developpement", 0, 0,
     "Etudes Geotech validees", "", "", "TBC", "2027-03-21", "",
     "Redimensionnement puissances besoins reels", "Etude Geotech BRG 100%", "Paiement solde 50% BRG; Etude offres leasing Ithemba", ""],
    ["tamatave", "Lidera Tamatave Phase 2", "Aymar", "Solaire", 16, "Developpement", 0, 0,
     "Etudes Geotech en cours", "", "", "TBC", "2027-03-24", "",
     "Discussion terrain en cours (Immo)", "Contrat etude geotech signe", "Relance Immo terrain; Paiement stand-by attente foncier", ""],
    ["oursun-1", "OurSun ZFI Phase 1", "Jordy", "Solaire", 3.2, "Developpement", 10, 0,
     "70% - Livrable provisoire recu", "", "", "BLUESKY", "2027-01-30", "",
     "Loyer IMMO impaye 51164 USD (Jan-Dec 2025)", "Etude faisabilite Artelia 70%", "Debut travaux 09/06/2026; HSE preparation; Structuration partenariat BlueSky", ""],
    ["oursun-2", "OurSun ZFI Phase 2", "Jordy", "Solaire", 7.8, "Planifie", 0, 0,
     "", "", "", "TBC", "2027-11-20", "",
     "Renforcement toiture en cours", "", "Attente toiture renforcee", ""],
    ["bongatsara-2", "Bongatsara Phase 2", "Henintsoa", "Solaire", 5, "Planifie", 0, 0,
     "", "", "", "TBC", "2027-08-12", "",
     "Changement terrain vers Ambohijanaka", "", "Terrain pas encore defini", ""],
    ["fihaonana-1", "Vestop Fihaonana Phase 1", "Toki", "Solaire", 4, "Developpement", 24, 0,
     "Faisabilite Artelia 90%", "", "Livrable Enermad en verification", "TBC", "2026-12-15", "",
     "Attente offre ligne evacuation JIRAMA", "Consultation EPC lancee DL 27/03", "Depot permis env.; Paiement Geo-Eco 70% restant; Validation controle SOCOTEC", ""],
    ["fihaonana-2", "Vestop Fihaonana Phase 2", "Toki", "Solaire", 15, "Planifie", 0, 0,
     "", "", "", "TBC", "2027-06-15", "",
     "", "", "Attente offres etude faisabilite + geotech; DL validation env. 18/03", ""],
    ["diego-wind-2", "Diego Wind Phase 2", "Toki", "Eolien", 4.88, "Planifie", 0, 0,
     "Consultation lancee", "", "", "TBC", "2027-09-15", "",
     "", "", "Consultation etude faisabilite; Etude Topo debut avril", ""],
    ["moramanga-2", "Moramanga Phase 2 (MSM)", "Lindon", "Solaire", 25, "Planifie", 0, 0,
     "35%", "", "Fiche tri en attente plan foncier", "TBC", "2027-05-30", "",
     "", "", "Debut travaux 30/06/2026; CDC a mettre a jour pour consultation EPC", ""],
    ["nosy-be-2", "Nosy-Be Phase 2", "Lindon", "Solaire + BESS", 2, "Developpement", 0, 0,
     "90% - Etudes completees", "", "Permis construction attente signature EPC", "TBC", "2026-11-30", "",
     "Paiement permis en attente signature contrat EPC", "", "Offres EPC PV+BESS recues; Debut travaux 01/05/26", ""],
    ["tulear-3", "Tulear Phase 3", "Toki", "Solaire + BESS", 3.1, "Developpement", 0, 0,
     "100% - Etude effectuee", "", "", "TBC", "2027-01-23", "",
     "", "Comparatif offres EPC en cours", "Appel offres EPC (7 soumissionnaires); Validation SOCOTEC", ""],
    ["marais-masay", "Floating Solar Marais Masay", "Aymar", "Flottant", 10, "Planifie", 0, 0,
     "En cours", "", "", "TBC", "2027-11-10", "",
     "Devis BRG en attente validation DG depuis 27/01", "Envoi docs techniques NTPC par Sayouba", "Acquisition lac - procedure fonciere; Confirmation JIRAMA ligne evacuation", ""],
    ["ambohidratrimo", "Ambohidratrimo", "Toki", "Solaire", 10, "Planifie", 0, 0,
     "", "", "", "TBC", "2027-10-07", "",
     "Changement terrain", "", "Validation terrain Ambohimalaza pour 5MW", ""],
    ["small-sites", "Small Sites (4 sites)", "Aymar", "Solaire", 2.18, "Planifie", 0, 0,
     "Etudes preliminaires", "", "", "OTC/BLUESKY", "2027-07-20", "",
     "Terrains indisponibles Ihosy/Ranohira", "", "Offres leasing recues 3 sites; Consultation BlueSky envoyee 27/02; Confirmation descente sites", ""],
    ["lidera-toamasina-3", "Lidera Toamasina Phase 3", "Aymar", "Solaire", 2, "Planifie", 0, 0,
     "0%", "", "", "TBC", "2027-09-24", "",
     "", "", "", ""],
]

BASE_DIR = r"C:\Users\Meril\OneDrive - GROUPE FILATEX\Bureau\Fichiers de DOSSIER DASHBOARD - Data_Dashbords\01_Energy\Projet\EnR"
os.makedirs(BASE_DIR, exist_ok=True)

# =============================================
# FILE 1: AVANCEMENT PROJETS
# =============================================
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Avancement Projets"
ws1.sheet_properties.tabColor = GREEN

ws1.merge_cells("A1:R1")
ws1["A1"] = "Weekly Report EnR - Avancement Projets"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws1["A1"].fill = PatternFill("solid", fgColor=DARK_BG)

ws1.merge_cells("A2:D2")
ws1["A2"] = "Semaine du:"
ws1["A2"].font = subtitle_font
ws1["A2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws1["E2"] = date(2026, 3, 16)
ws1["E2"].font = green_font
ws1["E2"].number_format = "DD/MM/YYYY"
ws1["E2"].fill = PatternFill("solid", fgColor=DARK_BG)

ws1.merge_cells("G2:H2")
ws1["G2"] = "Compile par:"
ws1["G2"].font = subtitle_font
ws1["G2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws1["I2"] = "PMO"
ws1["I2"].font = green_font
ws1["I2"].fill = PatternFill("solid", fgColor=DARK_BG)

ws1.merge_cells("K2:L2")
ws1["K2"] = "Prochaine reunion:"
ws1["K2"].font = subtitle_font
ws1["K2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws1["M2"] = date(2026, 3, 23)
ws1["M2"].font = green_font
ws1["M2"].number_format = "DD/MM/YYYY"
ws1["M2"].fill = PatternFill("solid", fgColor=DARK_BG)

headers = [
    "ID Projet", "Projet", "Responsable", "Type", "Puissance (MWc)",
    "Phase", "Avancement (%)", "Glissement (jours)",
    "Statut Engineering", "Statut Construction", "Statut Permis",
    "EPC / Contractant", "Date MES prevue", "Date MES revisee",
    "Blocages & Risques", "Actions S-1 (realisees)", "Actions S (a faire)", "Commentaires DG"
]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header(ws1, 4, len(headers))

for idx, proj in enumerate(projects):
    row = 5 + idx
    for col, val in enumerate(proj, 1):
        ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, len(headers), alt=(idx % 2 == 1))

col_widths = [14, 28, 12, 10, 12, 14, 14, 14, 20, 22, 18, 16, 14, 14, 35, 35, 35, 25]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "C5"

# Config sheet in avancement file
ws_cfg = wb1.create_sheet("Config")
ws_cfg.sheet_properties.tabColor = CYAN

ws_cfg.merge_cells("A1:D1")
ws_cfg["A1"] = "Configuration - Mapping Dashboard"
ws_cfg["A1"].font = Font(name="Arial", bold=True, size=12, color=CYAN)
ws_cfg["A1"].fill = PatternFill("solid", fgColor=DARK_BG)

ws_cfg["A2"] = "Ne pas modifier cet onglet - utilise par le pipeline automatique"
ws_cfg["A2"].font = Font(name="Arial", italic=True, size=9, color=RED)
ws_cfg["A2"].fill = PatternFill("solid", fgColor=DARK_BG)

config_headers = ["ID Dashboard", "Nom Projet", "Responsable", "Phase initiale"]
for i, h in enumerate(config_headers, 1):
    ws_cfg.cell(row=4, column=i, value=h)
style_header(ws_cfg, 4, len(config_headers))

for idx, proj in enumerate(projects):
    row = 5 + idx
    ws_cfg.cell(row=row, column=1, value=proj[0])
    ws_cfg.cell(row=row, column=2, value=proj[1])
    ws_cfg.cell(row=row, column=3, value=proj[2])
    ws_cfg.cell(row=row, column=4, value=proj[5])
    style_data_row(ws_cfg, row, 4, alt=(idx % 2 == 1))

config_widths = [20, 30, 14, 16]
for i, w in enumerate(config_widths, 1):
    ws_cfg.column_dimensions[get_column_letter(i)].width = w
ws_cfg.freeze_panes = "A5"

for ws in wb1.worksheets:
    apply_dark_bg(ws)

out1 = os.path.join(BASE_DIR, "Weekly_EnR_Avancement.xlsx")
wb1.save(out1)
print(f"FILE 1 OK: {out1}")
print(f"  Sheets: {wb1.sheetnames}")
print(f"  Projects: {len(projects)}")

# =============================================
# FILE 2: PAIEMENTS & FACTURES
# =============================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Paiements"
ws2.sheet_properties.tabColor = ORANGE

ws2.merge_cells("A1:J1")
ws2["A1"] = "Suivi Paiements EnR"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color=ORANGE)
ws2["A1"].fill = PatternFill("solid", fgColor=DARK_BG)

ws2.merge_cells("A2:D2")
ws2["A2"] = "Semaine du:"
ws2["A2"].font = subtitle_font
ws2["A2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws2["E2"] = date(2026, 3, 16)
ws2["E2"].font = green_font
ws2["E2"].number_format = "DD/MM/YYYY"
ws2["E2"].fill = PatternFill("solid", fgColor=DARK_BG)

pay_headers = [
    "ID Projet", "Projet", "Contractant", "Prestation", "Description",
    "Montant", "Devise", "Echeance", "Statut", "Actions en cours"
]
for i, h in enumerate(pay_headers, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(pay_headers))

payments = [
    ["diego-wind-1", "Diego Wind Ph1", "AEOLOS", "2nd livraison maritime", "2e shipment", 15600, "USD", "2026-01-15", "OK", ""],
    ["diego-wind-1", "Diego Wind Ph1", "AEOLOS", "EPC Materiels", "50% Manufacturing", 68317, "USD", "", "OK", ""],
    ["diego-wind-1", "Diego Wind Ph1", "AEOLOS", "EPC Materiels", "Meter purchase", 5000, "USD", "2025-09-22", "NOK", "OV signe, attente nivellement"],
    ["diego-wind-1", "Diego Wind Ph1", "AEOLOS", "EPC Materiels", "Before shipping rest 5 turbines", 85396, "USD", "2025-12-09", "NOK", "OV signe, attente nivellement"],
    ["diego-wind-1", "Diego Wind Ph1", "OTI", "Installation eolien", "3e paiement", 24619, "USD", "2025-11-15", "NOK", "OV en cours de preparation"],
    ["diego-wind-1", "Diego Wind Ph1", "OTI", "Installation eolien", "4e paiement", 24619, "USD", "2025-12-15", "NOK", "Facture receptionnee"],
    ["diego-lidera", "Lidera Diego", "LAGEOTECH", "Etude geotech", "Acompte demarrage", 953, "USD", "2025-12-02", "NOK", "OV en cours de signature"],
    ["mahajanga", "Lidera Majunga", "BRG", "Etude geotech", "Acompte 50%", 2105, "USD", "2025-11-02", "NOK", "Facture receptionnee"],
    ["fihaonana-1", "Vestop Fihaonana", "ARTELIA", "Etude faisabilite", "30% avance signature", 32821000, "MGA", "2026-01-04", "NOK", "Paiement a preparer"],
    ["fihaonana-1", "Vestop Fihaonana", "GEO ECO", "Etude geotech", "Solde 70%", 8019144, "MGA", "2026-01-17", "NOK", "Paiement a preparer"],
    ["bongatsara-1", "Bongatsara Ph1", "CERMAD", "GC", "25% avancement 50% travaux", 304913805, "MGA", "2026-03-18", "NOK", "ARRET TRAVAUX retard paiement"],
    ["oursun-1", "OurSun ZFI Ph1", "ARTELIA", "Etude faisabilite", "40% remise livrable", 16542762, "MGA", "2026-03-06", "NOK", ""],
    ["moramanga-1", "Moramanga Ph1", "MIDEX", "Dedouanement", "Batch modules", 293405798, "MGA", "2026-04-03", "NOK", ""],
]

for idx, pay in enumerate(payments):
    row = 5 + idx
    for col, val in enumerate(pay, 1):
        ws2.cell(row=row, column=col, value=val)
    style_data_row(ws2, row, len(pay_headers), alt=(idx % 2 == 1))

# Summary row
summary_row = 5 + len(payments) + 1
ws2.cell(row=summary_row, column=5, value="TOTAL USD").font = Font(name="Arial", bold=True, size=9, color=ORANGE)
ws2.cell(row=summary_row, column=6).value = f"=SUMPRODUCT((G5:G{4+len(payments)}=\"USD\")*F5:F{4+len(payments)})"
ws2.cell(row=summary_row, column=6).font = Font(name="Arial", bold=True, size=9, color=ORANGE)
ws2.cell(row=summary_row, column=6).number_format = "#,##0"
style_data_row(ws2, summary_row, len(pay_headers))

summary_row2 = summary_row + 1
ws2.cell(row=summary_row2, column=5, value="TOTAL MGA").font = Font(name="Arial", bold=True, size=9, color=ORANGE)
ws2.cell(row=summary_row2, column=6).value = f"=SUMPRODUCT((G5:G{4+len(payments)}=\"MGA\")*F5:F{4+len(payments)})"
ws2.cell(row=summary_row2, column=6).font = Font(name="Arial", bold=True, size=9, color=ORANGE)
ws2.cell(row=summary_row2, column=6).number_format = "#,##0"
style_data_row(ws2, summary_row2, len(pay_headers))

pay_widths = [14, 20, 16, 20, 30, 16, 8, 14, 8, 35]
for i, w in enumerate(pay_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "C5"

# Factures sheet
ws3 = wb2.create_sheet("Factures")
ws3.sheet_properties.tabColor = RED

ws3.merge_cells("A1:I1")
ws3["A1"] = "Suivi Factures EnR"
ws3["A1"].font = Font(name="Arial", bold=True, size=14, color=RED)
ws3["A1"].fill = PatternFill("solid", fgColor=DARK_BG)

ws3.merge_cells("A2:D2")
ws3["A2"] = "Semaine du:"
ws3["A2"].font = subtitle_font
ws3["A2"].fill = PatternFill("solid", fgColor=DARK_BG)
ws3["E2"] = date(2026, 3, 16)
ws3["E2"].font = green_font
ws3["E2"].number_format = "DD/MM/YYYY"
ws3["E2"].fill = PatternFill("solid", fgColor=DARK_BG)

fact_headers = [
    "Ref Facture", "Montant", "Devise", "Fournisseur",
    "Date Facture", "Echeance", "Projet", "Statut", "Commentaire"
]
for i, h in enumerate(fact_headers, 1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(fact_headers))

fact_widths = [20, 16, 8, 16, 14, 14, 20, 10, 40]
for i, w in enumerate(fact_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A5"

for ws in wb2.worksheets:
    apply_dark_bg(ws)

out2 = os.path.join(BASE_DIR, "Weekly_EnR_Paiements.xlsx")
wb2.save(out2)
print(f"FILE 2 OK: {out2}")
print(f"  Sheets: {wb2.sheetnames}")
print(f"  Payments: {len(payments)}")
