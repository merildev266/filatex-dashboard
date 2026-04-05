"""
Generate enr_projects_data.js from all EnR project Excel files.
Replaces the hardcoded enrProjects + enrEnrich in energy.js.

Sources (01_Energy/Projet/EnR/):
  - ENR_Dates_Projets_Budgets.xlsx  → base project data (dates, capex, MWc, TRI, quarters)
  - ENR_Cost_Control.xlsx           → cost control (BAC, AC, SPI, CPI, performance)
  - ENR_Dashboard_Master.xlsx       → DATABASE (lead, epc, glissement, puissance, prodJour)
                                      PROJECT STATUS (constProg)
  - ENR_Status_Etudes.xlsx          → DEVELOPMENT (study progress per category)
  - Weekly_EnR_Avancement.xlsx      → weekly overrides (avancement, blocages, actions)
  - Weekly_EnR_Paiements.xlsx       → payment tracking (montants, statuts, alertes)
"""
import json, os, sys, io
from datetime import datetime, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

BASE = os.path.join(
    os.environ.get("USERPROFILE", r"C:\Users\Meril"),
    "OneDrive - GROUPE FILATEX",
    "Bureau",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "01_Energy", "Projet", "EnR",
)

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "enr_projects_data.js")

# ══ PROJECT ID MAPPING ══
# Maps various Excel names to dashboard IDs
NAME_MAP = {
    "nosy-be phase 1": "nosy-be-1",
    "nosy-be phase1 3mwc": "nosy-be-1",
    "nosy be 3mw": "nosy-be-1",
    "tulear phase 2": "tulear-2",
    "tulear phase2 1mwc": "tulear-2",
    "tulear ground 1mw": "tulear-2",
    "moramanga phase 1": "moramanga-1",
    "mandrosolar phase1 15mwc": "moramanga-1",
    "msm ph1 15mw": "moramanga-1",
    "top energie diego wind phase 1": "diego-wind-1",
    "te diego wind phase1 0.12mwc": "diego-wind-1",
    "diego wind 120 kw": "diego-wind-1",
    "top energie ground solar bongatsara phase 1": "bongatsara-1",
    "te bongatsara phase1 5mwc": "bongatsara-1",
    "bongatsara ph1 5mw": "bongatsara-1",
    "bongatsara ph1: 5mw": "bongatsara-1",
    "oursun zfi phase 1": "oursun-1",
    "oursun rooftop phase1 3.2mwc": "oursun-1",
    "oursun 3,2 mw": "oursun-1",
    "vestop fihaonana phase 1": "fihaonana-1",
    "vestop fihaonana phase1 4mwc": "fihaonana-1",
    "vestop fihaonana 4mw": "fihaonana-1",
    "nosy-be phase 2": "nosy-be-2",
    "nosy-be phase2 2mwc": "nosy-be-2",
    "tulear phase 3": "tulear-3",
    "tulear phase3 3.1mwc": "tulear-3",
    "moramanga phase 2": "moramanga-2",
    "mandrosolar phase2 25mwc": "moramanga-2",
    "oursun zfi phase 2": "oursun-2",
    "oursun rooftop phase2 7.8mwc": "oursun-2",
    "top energie ground solar bongatsara phase 2": "bongatsara-2",
    "te bongatsara phase2 5mwc": "bongatsara-2",
    "top energie diego wind phase 2": "diego-wind-2",
    "te diego wind phase2 4.88mwc": "diego-wind-2",
    "vestop fihaonana phase 2": "fihaonana-2",
    "vestop fihaonana phase2 15mwc": "fihaonana-2",
    "top energie floating solar marais masay": "marais-masay",
    "te marais masay 10mwc floating": "marais-masay",
    "top energie ambohidratrimo": "ambohidratrimo",
    "te ambohidratrimo 10mwc": "ambohidratrimo",
    "small sites": "small-sites",
    "filatex energie small sites 2.175mwc": "small-sites",
    "lidera tamatave": "tamatave",
    "lidera phase2 toamasina 16mwc": "tamatave",
    "tamatave phase 2": "tamatave",
    "lidera diego": "diego-lidera",
    "lidera phase2 diego 7.6mwc": "diego-lidera",
    "diego phase 2": "diego-lidera",
    "lidera mahajanga": "mahajanga",
    "lidera phase2 majunga 10.75mwc": "mahajanga",
    "majunga phase 2": "mahajanga",
    "lidera phase3 toamasina 2mwc": "lidera-toamasina-3",
    "lidera toamasina phase 3": "lidera-toamasina-3",
    # Paiements sheet names
    "diego wind ph1": "diego-wind-1",
    "lidera diego": "diego-lidera",
    "lidera majunga": "mahajanga",
    "vestop fihaonana": "fihaonana-1",
    "bongatsara ph1": "bongatsara-1",
    "oursun zfi ph1": "oursun-1",
    "moramanga ph1": "moramanga-1",
    # Cost Control RECAP names
    "bongatsara ph1 5mw": "bongatsara-1",
    "tulear ground 1mw": "tulear-2",
    "msm ph1 15mw": "moramanga-1",
    "majunga phase 2": "mahajanga",
    "diego phase 2": "diego-lidera",
    "tamatave phase 2": "tamatave",
}


def _normalize(s):
    """Normalize string for matching: lowercase, strip accents, collapse spaces."""
    import unicodedata
    s = s.strip().lower()
    s = s.rstrip('#').rstrip().rstrip(',').strip()
    # Remove accents (é→e, etc.)
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    # Collapse multiple spaces
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s


# Pre-build normalized NAME_MAP
_NORM_MAP = {_normalize(k): v for k, v in NAME_MAP.items()}


def resolve_id(name):
    if not name:
        return None
    n = _normalize(name)
    # Direct match
    if n in _NORM_MAP:
        return _NORM_MAP[n]
    # Partial match
    for key, pid in _NORM_MAP.items():
        if key in n or n in key:
            return pid
    return None


def safe_float(v, default=None):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 10 and s[4] == '-':
        return s[:10]
    return None


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


# ══════════════════════════════════════════
# FILE 1: ENR_Dates_Projets_Budgets.xlsx
# ══════════════════════════════════════════
def read_dates_budgets(projects):
    path = os.path.join(BASE, "ENR_Dates_Projets_Budgets.xlsx")
    if not os.path.exists(path):
        print(f"  SKIP: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Master Plan"]

    for row in range(4, ws.max_row + 1):
        name = safe_str(ws.cell(row, 2).value)
        # Stop at TOTAL row — only read the first section (Master Plan)
        if name.upper().startswith("TOTAL"):
            break
        pid = resolve_id(name)
        if not pid:
            continue

        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        p["name"] = name.rstrip('#').strip()
        p["pvMw"] = safe_float(ws.cell(row, 3).value)
        p["bessMwh"] = safe_float(ws.cell(row, 4).value)
        p["capexM"] = safe_float(ws.cell(row, 9).value)
        tri_val = safe_float(ws.cell(row, 10).value)
        if tri_val is not None:
            p["tri"] = round(tri_val * 100, 1) if tri_val < 1 else round(tri_val, 1)
        p["engStart"] = safe_date(ws.cell(row, 11).value)
        p["engEnd"] = safe_date(ws.cell(row, 12).value)
        eng_pct = ws.cell(row, 13).value
        if eng_pct is not None:
            p["engPct"] = round(safe_float(eng_pct, 0) * 100) if safe_float(eng_pct, 0) <= 1 else round(safe_float(eng_pct, 0))
        tend_val = ws.cell(row, 14).value
        if tend_val and str(tend_val).strip().lower() == "done":
            p["tendDone"] = True
            p["tendStart"] = None
        else:
            p["tendStart"] = safe_date(tend_val)
            p["tendDone"] = False
        p["tendEnd"] = safe_date(ws.cell(row, 15).value)
        p["constStart"] = safe_date(ws.cell(row, 16).value)
        p["constEnd"] = safe_date(ws.cell(row, 17).value)
        p["costDev"] = safe_float(ws.cell(row, 18).value)
        p["costPv"] = safe_float(ws.cell(row, 19).value)

        # Quarterly funds (cols 20-29 = Q3-25 to Q4-27)
        qtr_labels = ["Q3-25","Q4-25","Q1-26","Q2-26","Q3-26","Q4-26","Q1-27","Q2-27","Q3-27","Q4-27"]
        qtr = []
        for i, ql in enumerate(qtr_labels):
            val = safe_float(ws.cell(row, 20 + i).value)
            if val and val > 0:
                qtr.append({"q": ql, "a": round(val)})
        if qtr:
            p["qtr"] = qtr

        p["comment"] = safe_str(ws.cell(row, 30).value) or None

    wb.close()
    print(f"  Dates/Budgets: {len([p for p in projects.values() if 'pvMw' in p])} projects")


# ══════════════════════════════════════════
# FILE 2: ENR_Cost_Control.xlsx
# ══════════════════════════════════════════
def read_cost_control(projects):
    path = os.path.join(BASE, "ENR_Cost_Control.xlsx")
    if not os.path.exists(path):
        print(f"  SKIP: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["RECAP"]

    count = 0
    for row in range(5, ws.max_row + 1):
        name = safe_str(ws.cell(row, 1).value)
        pid = resolve_id(name)
        if not pid:
            continue
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        bac = safe_float(ws.cell(row, 3).value)
        forecast = safe_float(ws.cell(row, 4).value)
        ac = safe_float(ws.cell(row, 7).value)
        av_reel = ws.cell(row, 8).value
        spi = safe_float(ws.cell(row, 13).value)
        cpi = safe_float(ws.cell(row, 17).value)
        perf = safe_str(ws.cell(row, 16).value)

        # Parse avancement reel (can be decimal, integer, or text)
        av_val = None
        if av_reel is not None:
            f = safe_float(av_reel)
            if f is not None:
                av_val = round(f * 100) if f <= 1 else round(f)

        cc = {}
        if bac is not None: cc["bac"] = round(bac)
        if forecast is not None: cc["forecast"] = round(forecast)
        if ac is not None: cc["ac"] = round(ac)
        if av_val is not None: cc["avReel"] = av_val
        if spi is not None and spi < 100: cc["spi"] = round(spi, 2)
        if cpi is not None and cpi < 100: cc["cpi"] = round(cpi, 2)
        if perf: cc["perf"] = perf

        if cc:
            p["cc"] = cc
            count += 1

    wb.close()
    print(f"  Cost Control: {count} projects")


# ══════════════════════════════════════════
# FILE 3: ENR_Dashboard_Master.xlsx
# ══════════════════════════════════════════
def read_dashboard_master(projects):
    path = os.path.join(BASE, "ENR_Dashboard_Master.xlsx")
    if not os.path.exists(path):
        print(f"  SKIP: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)

    # DATABASE sheet
    ws_db = wb["DATABASE"]
    count_db = 0
    for row in range(3, ws_db.max_row + 1):
        name = safe_str(ws_db.cell(row, 2).value)
        pid = resolve_id(name)
        if not pid:
            continue
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        loc = safe_str(ws_db.cell(row, 3).value)
        lead = safe_str(ws_db.cell(row, 6).value)
        epciste = safe_str(ws_db.cell(row, 7).value)
        puissance = safe_float(ws_db.cell(row, 36).value)
        prod_jour = safe_float(ws_db.cell(row, 37).value)
        glissement = safe_float(ws_db.cell(row, 30).value)

        if loc: p["loc"] = loc
        if lead: p["lead"] = lead
        if epciste: p["epciste"] = epciste
        if puissance: p["puissance"] = round(puissance)
        if prod_jour: p["prodJour"] = round(prod_jour)
        if glissement is not None: p["glissement"] = int(glissement)
        count_db += 1

    # PROJECT STATUS sheet
    ws_ps = wb["PROJECT STATUS"]
    count_ps = 0
    for row in range(4, ws_ps.max_row + 1):
        name = safe_str(ws_ps.cell(row, 2).value)
        pid = resolve_id(name)
        if not pid:
            continue
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        stage = safe_str(ws_ps.cell(row, 4).value)
        const_prog = safe_float(ws_ps.cell(row, 11).value)

        if stage:
            if "construction" in stage.lower():
                p["phase"] = "Construction"
            elif "development" in stage.lower():
                p["phase"] = "Developpement"

        if const_prog is not None:
            p["constProg"] = round(const_prog, 2) if const_prog <= 1 else round(const_prog / 100, 2)
            count_ps += 1

    wb.close()
    print(f"  Dashboard Master: {count_db} (DB) + {count_ps} (STATUS)")


# ══════════════════════════════════════════
# FILE 4: ENR_Status_Etudes.xlsx
# ══════════════════════════════════════════
def read_status_etudes(projects):
    path = os.path.join(BASE, "ENR_Status_Etudes.xlsx")
    if not os.path.exists(path):
        print(f"  SKIP: {path}")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["DEVELOPMENT"]

    study_names = [
        "Etude Topographique", "Etude Geotechnique", "Test Arrachement",
        "Etude Faisabilite", "Etude de Ligne", "Categorisation Env.",
        "Etude Environnementale", "Permis Environnemental", "Permis Construire",
        "Autorisation ACM", "Etude Technico-Financiere"
    ]

    count = 0
    for row in range(3, ws.max_row + 1):
        name = safe_str(ws.cell(row, 1).value)
        pid = resolve_id(name)
        if not pid:
            continue
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        # Read study progress (columns 2-12)
        studies = []
        for i, sn in enumerate(study_names):
            val = safe_float(ws.cell(row, 2 + i).value)
            pct = round(val) if val is not None else 0
            studies.append({"name": sn, "pct": pct})

        # Progression M (col 15)
        prog_m = safe_float(ws.cell(row, 15).value)
        if prog_m is not None:
            p["engProgression"] = round(prog_m * 100) if prog_m <= 1 else round(prog_m)

        if any(s["pct"] > 0 for s in studies):
            p["studies"] = studies
            count += 1

    wb.close()
    print(f"  Status Etudes: {count} projects")


# ══════════════════════════════════════════
# FILE 5: Weekly_EnR_Avancement.xlsx
# New format: 52 sheets S01-S52, auto-detect latest filled week
# Headers row 8: A=ID, B=Projet, C=Resp, D=Type, E=MWc,
#   F=Phase, G=Avancement(%), H=Glissement(j), I=StatutEtudes,
#   J=StatutConstruction, K=EPC, L=COD, M=Blocages,
#   N=ActionsS-1, O=ActionsSemaine, P=CommentairesDG
# ══════════════════════════════════════════
def _find_latest_weekly_sheet(wb):
    """Find the latest sheet with actual data (non-empty F column)."""
    best = None
    for sn in reversed(wb.sheetnames):
        if not sn.startswith('S') or not sn[1:].isdigit():
            continue
        ws = wb[sn]
        # Check if any project row has Phase filled (col F, rows 9+)
        for row in range(9, 30):
            if ws.cell(row, 6).value:
                return sn
    return None


def read_weekly(projects):
    path = os.path.join(BASE, "Weekly_Report", "Weekly_EnR_Avancement.xlsx")
    if not os.path.exists(path):
        path = os.path.join(BASE, "Weekly_EnR_Avancement.xlsx")
    if not os.path.exists(path):
        print(f"  SKIP: {path}")
        return None
    wb = openpyxl.load_workbook(path, data_only=True)

    # Try new multi-sheet format first
    latest = _find_latest_weekly_sheet(wb)
    if latest:
        ws = wb[latest]
        data_start = 9  # row 9 = first data row in new format
        print(f"  Weekly format: multi-sheet, reading {latest}")
    elif "Avancement Projets" in wb.sheetnames:
        ws = wb["Avancement Projets"]
        data_start = 5  # old format
        latest = "old"
        print(f"  Weekly format: legacy single-sheet")
    else:
        print(f"  SKIP: no valid sheet found")
        wb.close()
        return None

    week_val = ws["E2"].value or ws["E3"].value
    week_str = ""
    if week_val:
        week_str = week_val.strftime("%d/%m/%Y") if hasattr(week_val, 'strftime') else str(week_val)

    # Column mapping depends on format
    if latest != "old":
        # New format columns: F=Phase G=Av H=Glis I=StatEtudes J=StatConst K=EPC L=COD M=Bloc N=ActS1 O=ActS P=ComDG
        col_map = {'resp':3, 'type':4, 'phase':6, 'av':7, 'glis':8,
                   'statEtudes':9, 'statConst':10, 'epc':11, 'cod':12,
                   'blocages':13, 'actionsS1':14, 'actionsS':15, 'comDG':16}
    else:
        # Old format columns
        col_map = {'resp':3, 'type':4, 'phase':6, 'av':7, 'glis':8,
                   'statEtudes':9, 'statConst':10, 'epc':12, 'cod':13,
                   'blocages':15, 'actionsS1':16, 'actionsS':17, 'comDG':18}

    count = 0
    for row in range(data_start, ws.max_row + 1):
        pid = safe_str(ws.cell(row, 1).value)
        if not pid:
            continue
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]

        resp = safe_str(ws.cell(row, col_map['resp']).value)
        type_ = safe_str(ws.cell(row, col_map['type']).value)
        phase = safe_str(ws.cell(row, col_map['phase']).value)
        avancement = safe_float(ws.cell(row, col_map['av']).value)
        glissement = safe_float(ws.cell(row, col_map['glis']).value)
        stat_etudes = safe_str(ws.cell(row, col_map['statEtudes']).value)
        stat_const = safe_str(ws.cell(row, col_map['statConst']).value)
        epc = safe_str(ws.cell(row, col_map['epc']).value)
        cod = safe_date(ws.cell(row, col_map['cod']).value)
        blocages = safe_str(ws.cell(row, col_map['blocages']).value)
        actions_s1 = safe_str(ws.cell(row, col_map['actionsS1']).value)
        actions_s = safe_str(ws.cell(row, col_map['actionsS']).value)
        commentaires_dg = safe_str(ws.cell(row, col_map['comDG']).value)

        if resp: p["chef"] = resp
        if type_: p["type"] = type_
        if phase: p["phase"] = phase
        if avancement is not None:
            p["avancement"] = round(avancement)
            p["constProg"] = round(avancement / 100, 2)
        if glissement is not None: p["glissement"] = int(glissement)
        if stat_etudes: p["statutEtudes"] = stat_etudes
        if stat_const: p["statutConst"] = stat_const
        if epc: p["epciste"] = epc
        if cod: p["codPrevue"] = cod
        if blocages: p["blocages"] = blocages
        if actions_s1: p["actionsS1"] = actions_s1
        if actions_s: p["actionsS"] = actions_s
        if commentaires_dg: p["commentairesDG"] = commentaires_dg
        count += 1

    wb.close()
    print(f"  Weekly: {count} projects (week: {week_str})")
    return week_str


# ══════════════════════════════════════════
# FILE 6: Weekly_EnR_Paiements.xlsx
# ══════════════════════════════════════════
def read_paiements(projects):
    # Try PAIEMENT sheet inside Weekly_EnR_Avancement first (new unified format)
    path = os.path.join(BASE, "Weekly_Report", "Weekly_EnR_Avancement.xlsx")
    if not os.path.exists(path):
        path = os.path.join(BASE, "Weekly_EnR_Avancement.xlsx")
    wb = None
    ws = None
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        if "PAIEMENT" in wb.sheetnames:
            ws = wb["PAIEMENT"]
    # Fallback: separate Weekly_EnR_Paiements.xlsx
    if ws is None:
        if wb: wb.close()
        wb = None
        for fname in ["Weekly_EnR_Paiements.xlsx"]:
            for base in [os.path.join(BASE, "Weekly_Report"), BASE]:
                fpath = os.path.join(base, fname)
                if os.path.exists(fpath):
                    try:
                        wb = openpyxl.load_workbook(fpath, data_only=True)
                        for sn in ["Paiements", "PAIEMENT", "Paiement"]:
                            if sn in wb.sheetnames:
                                ws = wb[sn]
                                break
                        if ws: break
                        wb.close(); wb = None
                    except Exception:
                        pass
            if ws: break
        if ws is None:
            print(f"  SKIP Paiements: no valid sheet found")
            if wb: wb.close()
            return

    # Detect format: PAIEMENT sheet starts at B2 (new) or A4 (old Paiements)
    # New: B=PROJET C=CONTRACTANT D=PRESTATION E=DESCRIPTION F=PROPORTION G=Montant H=Echeance I=Statut J=Actions
    # Old: A=ID B=Projet C=Contractant D=Prestation E=Description F=Montant G=Devise H=Echeance I=Statut J=Actions
    is_new_fmt = ws.title == "PAIEMENT"
    data_start = 3 if is_new_fmt else 5
    col_offset = 1 if is_new_fmt else 0  # new format starts at B, old at A

    pmt_data = {}

    for row in range(data_start, ws.max_row + 1):
        projet_name = safe_str(ws.cell(row, 2 + col_offset).value)
        if not projet_name:
            continue
        if "total" in projet_name.lower():
            continue

        pid = resolve_id(projet_name)
        if not pid:
            continue

        contractant = safe_str(ws.cell(row, 3 + col_offset).value)
        prestation = safe_str(ws.cell(row, 4 + col_offset).value)
        description = safe_str(ws.cell(row, 5 + col_offset).value)
        montant = safe_float(ws.cell(row, 7 if is_new_fmt else 6).value, 0)
        devise = "USD"  # new format is USD by default
        if not is_new_fmt:
            devise = safe_str(ws.cell(row, 7).value).upper()
        echeance = safe_date(ws.cell(row, 8 + col_offset).value)
        statut = safe_str(ws.cell(row, 9 + col_offset).value).upper()
        actions = safe_str(ws.cell(row, 10 + col_offset).value)

        if pid not in pmt_data:
            pmt_data[pid] = {"lines": [], "totalUSD": 0, "totalMGA": 0, "ok": 0, "nok": 0}
        d = pmt_data[pid]

        line = {}
        if contractant: line["contractant"] = contractant
        if prestation: line["prestation"] = prestation
        if description: line["desc"] = description
        if montant: line["montant"] = montant
        if devise: line["devise"] = devise
        if echeance: line["echeance"] = echeance
        if statut: line["statut"] = statut
        if actions: line["actions"] = actions
        d["lines"].append(line)

        if "USD" in devise:
            d["totalUSD"] += montant
        else:
            d["totalMGA"] += montant
        if statut == "OK":
            d["ok"] += 1
        else:
            d["nok"] += 1

    # Merge into projects
    count = 0
    for pid, d in pmt_data.items():
        if pid not in projects:
            projects[pid] = {"id": pid}
        p = projects[pid]
        pmt = {
            "count": len(d["lines"]),
            "ok": d["ok"],
            "nok": d["nok"],
        }
        if d["totalUSD"] > 0:
            pmt["totalUSD"] = round(d["totalUSD"], 2)
        if d["totalMGA"] > 0:
            pmt["totalMGA"] = round(d["totalMGA"])
        pmt["lines"] = d["lines"]
        p["paiements"] = pmt
        count += 1

    wb.close()
    print(f"  Paiements: {count} projects, {sum(len(d['lines']) for d in pmt_data.values())} lines")


# ══════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════
def determine_phase(p):
    """Determine project phase from available data."""
    if "phase" in p:
        return p["phase"]
    cp = p.get("constProg")
    if cp is not None and cp >= 0.98:
        return "Termine"
    if cp is not None and cp > 0:
        return "Construction"
    eng = p.get("engPct", 0)
    if eng and eng > 0:
        return "Developpement"
    return "Planifie"


def build_enr_projects():
    """Build and return ENR projects data dict (for API use). Returns {"ENR_PROJECTS_DATA": {...}}."""
    projects = {}
    read_dates_budgets(projects)
    read_cost_control(projects)
    read_dashboard_master(projects)
    read_status_etudes(projects)
    week_str = read_weekly(projects)
    read_paiements(projects)

    for pid, p in projects.items():
        p["id"] = pid
        if "phase" not in p:
            p["phase"] = determine_phase(p)
        if "name" not in p:
            p["name"] = pid

    phase_order = {"Termine": 0, "Construction": 1, "Developpement": 2, "Planifie": 3}
    project_list = sorted(projects.values(), key=lambda p: (phase_order.get(p.get("phase", ""), 9), p.get("name", "")))

    data = {
        "generated": datetime.now().isoformat(),
        "week": week_str or "",
        "source": "01_Energy/Projet/EnR/",
        "projects": project_list,
    }
    return {"ENR_PROJECTS_DATA": data}


def main():
    print("Generating EnR projects data...")
    print(f"Source: {BASE}\n")

    result = build_enr_projects()
    data = result["ENR_PROJECTS_DATA"]
    project_list = data["projects"]
    week_str = data["week"]

    js = "/* Auto-generated by generate_enr_projects.py — DO NOT EDIT */\n"
    js += f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    js += "window.ENR_PROJECTS_DATA = " + json.dumps(data, ensure_ascii=False, indent=2, default=str) + ";\n"

    with open(OUT, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"\nOK: {OUT}")
    print(f"  Projects: {len(project_list)}")
    print(f"  Week: {week_str}")
    phases = {}
    for p in project_list:
        ph = p.get("phase", "?")
        phases[ph] = phases.get(ph, 0) + 1
    print(f"  Phases: {phases}")


if __name__ == "__main__":
    main()
