"""
Reads EnR production Excel files (rapport de production + production elum)
and generates enr_site_data.js for the dashboard.

Source: OneDrive/.../01_Energy/Production/EnR/
  - DIE_rapport_de_production.xlsx
  - TMM_rapport de production.xlsx
  - MJN_rapport de production.xlsx
  - DIE_production_elum.xlsx  (optional, 10-min granular)
  - TMM_production_elum.xlsx  (optional)
  - MJN_production_elum.xlsx  (optional)
"""
import json
import os
from datetime import datetime, timedelta

import openpyxl

ENR_DIR = os.path.join(
    os.environ.get("USERPROFILE", "C:/Users/Meril"),
    "OneDrive - GROUPE FILATEX",
    "Fichiers de DOSSIER DASHBOARD - Data_Dashbords",
    "01_Energy", "Production", "EnR",
)

SITES = {
    "DIE": {
        "name": "Diego",
        "entity": "Diego Green Power",
        "centrale": "Centrale Solaire d'Ankorikahely",
        "loc": "Ankorikahely, Diego",
        "capacityKwc": 2400,
        "file": "DIE_rapport_de_production.xlsx",
    },
    "TMM": {
        "name": "Tamatave",
        "entity": "Toamasina Green Power",
        "centrale": "Centrale Solaire de la Verrerie",
        "loc": "La Verrerie, Tamatave",
        "capacityKwc": 2000,
        "file": "TMM_rapport de production.xlsx",
    },
    "MJN": {
        "name": "Majunga",
        "entity": "Majunga Green Power",
        "centrale": "Centrale Solaire d'Andranotakatra",
        "loc": "Andranotakatra, Majunga",
        "capacityKwc": 1200,
        "file": "MJN_rapport de production.xlsx",
    },
}

# Month name patterns (French) for sheet name matching
MONTH_PATTERNS = {
    1: ["janv", "janvier"],
    2: ["fev", "f\u00e9v", "fevrier", "f\u00e9vrier"],
    3: ["mars"],
    4: ["avr", "avril"],
    5: ["mai"],
    6: ["juin"],
    7: ["juil", "juillet"],
    8: ["aout", "ao\u00fbt"],
    9: ["sept", "septembre"],
    10: ["oct", "octobre"],
    11: ["nov", "novembre"],
    12: ["dec", "d\u00e9c", "d\u00e9cembre", "decembre"],
}


def match_month(sheet_name):
    """Return month number (1-12) from a sheet name, or None."""
    sn = sheet_name.lower().strip()
    for month_num, patterns in MONTH_PATTERNS.items():
        for pat in patterns:
            if pat in sn:
                return month_num
    return None


def parse_index_sheet(ws):
    """Parse an 'Index compteur' sheet.
    Returns list of daily dicts with: date, production_kwh, peak_kw, peak_hour,
    irradiance, coupling_start, coupling_end, coupling_hours.
    """
    rows = []
    # Find the header row (contains 'DATE' and 'INDEX')
    header_row = None
    date_col = None
    index_col = None
    prod_col = None
    peak_kw_col = None
    peak_hour_col = None
    irradiance_col = None
    coupling_start_col = None
    coupling_end_col = None
    coupling_dur_col = None

    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip().upper() == "DATE":
                header_row = r
                date_col = c
                break
        if header_row:
            break

    if not header_row:
        return rows

    # Find columns by scanning header area
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if not v:
            # Check row above for merged headers
            v_above = ws.cell(header_row - 1, c).value
            if v_above:
                v_str = str(v_above).strip().upper()
            else:
                continue
        else:
            v_str = str(v).strip().upper()

        if "INDEX" in v_str:
            index_col = c
        elif "PRODUCTION" in v_str and "KWH" in v_str:
            prod_col = c
        elif v_str == "KW":
            peak_kw_col = c
        elif "HEURE" in v_str:
            peak_hour_col = c
        elif "IRRADIANCE" in v_str or "KWH/M" in v_str or "WH/M" in v_str:
            irradiance_col = c

    # Check row above header for coupling columns
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            v_str = str(v).strip().lower()
            if "d\u00e9but" in v_str or "debut" in v_str:
                coupling_start_col = c
            elif "fin" == v_str:
                coupling_end_col = c

    # Also check the row between header labels
    r_labels = header_row - 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r_labels, c).value
        if v:
            v_str = str(v).strip().lower()
            if "couplage" == v_str.strip():
                coupling_start_col = c
            elif "d\u00e9couplage" in v_str or "decouplage" in v_str:
                coupling_end_col = c
            elif "dur\u00e9e" in v_str or "duree" in v_str:
                coupling_dur_col = c

    # Read data rows starting from header_row + 1 (skip the initial index row)
    first_index_row = header_row + 1
    prev_index = None
    # Check if first data row has only index (previous day baseline)
    v_first_date = ws.cell(first_index_row, date_col).value
    v_first_prod = ws.cell(first_index_row, prod_col).value if prod_col else None
    if v_first_date and v_first_prod is None and index_col:
        prev_index = ws.cell(first_index_row, index_col).value
        first_index_row += 1

    for r in range(first_index_row, ws.max_row + 1):
        date_val = ws.cell(r, date_col).value
        if not date_val:
            continue
        # Skip summary rows (TOTAL, Énergie, etc.)
        if isinstance(date_val, str):
            if not date_val.strip()[:4].isdigit():
                continue
            date_str = date_val.strip()[:10]
        elif isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            continue

        prod_kwh = None
        if prod_col:
            prod_kwh = ws.cell(r, prod_col).value
        if prod_kwh is None and index_col and prev_index is not None:
            cur_index = ws.cell(r, index_col).value
            if cur_index is not None and prev_index is not None:
                try:
                    prod_kwh = float(cur_index) - float(prev_index)
                except (ValueError, TypeError):
                    pass

        if index_col:
            cur_idx = ws.cell(r, index_col).value
            if cur_idx is not None:
                try:
                    prev_index = float(cur_idx)
                except (ValueError, TypeError):
                    pass

        if prod_kwh is None or (isinstance(prod_kwh, (int, float)) and prod_kwh <= 0):
            continue

        peak_kw = ws.cell(r, peak_kw_col).value if peak_kw_col else None
        irradiance = ws.cell(r, irradiance_col).value if irradiance_col else None

        row_data = {
            "date": date_str,
            "prodKwh": round(float(prod_kwh), 1) if prod_kwh else 0,
        }
        if peak_kw is not None:
            try:
                row_data["peakKw"] = round(float(peak_kw), 0)
            except (ValueError, TypeError):
                pass
        if irradiance is not None:
            try:
                row_data["irradiance"] = round(float(irradiance), 2)
            except (ValueError, TypeError):
                pass

        # Coupling hours
        if coupling_dur_col:
            dur_val = ws.cell(r, coupling_dur_col).value
            if dur_val is not None:
                try:
                    row_data["couplingHours"] = round(float(dur_val), 2)
                except (ValueError, TypeError):
                    pass

        rows.append(row_data)

    return rows


def parse_prod_solaire_sheet(ws):
    """Parse a 'Prod solaire' sheet.
    Returns list of daily dicts with: date, availability_hours, scheduled_interruptions,
    unscheduled_interruptions, capacity_kwc, energy_produced_kwh, energy_consumed_kwh, energy_delivered_kwh.
    """
    rows = []
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if v and str(v).strip().upper() == "DATE":
            header_row = r
            break

    if not header_row:
        return rows

    for r in range(header_row + 1, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        if not date_val:
            continue
        # Skip summary rows (TOTAL, Énergie, etc.)
        if isinstance(date_val, str):
            if not date_val.strip()[:4].isdigit():
                continue
            date_str = date_val.strip()[:10]
        elif isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            continue

        avail = ws.cell(r, 2).value
        sched_int = ws.cell(r, 3).value
        unsched_int = ws.cell(r, 4).value
        capacity = ws.cell(r, 5).value
        produced = ws.cell(r, 6).value
        consumed = ws.cell(r, 7).value
        delivered = ws.cell(r, 8).value

        if produced is None:
            continue

        row_data = {
            "date": date_str,
            "availHours": round(float(avail), 2) if avail else 0,
            "schedInterrupt": round(float(sched_int), 2) if sched_int else 0,
            "unschedInterrupt": round(float(unsched_int), 2) if unsched_int else 0,
            "capacityKwc": round(float(capacity)) if capacity else 0,
            "prodKwh": round(float(produced), 1) if produced else 0,
            "consumedKwh": round(float(consumed), 2) if consumed else 0,
            "deliveredKwh": round(float(delivered), 1) if delivered else 0,
        }
        rows.append(row_data)

    return rows


def build_site_data(code, site_cfg):
    """Build complete site data object from Excel file."""
    filepath = os.path.join(ENR_DIR, site_cfg["file"])
    if not os.path.exists(filepath):
        print(f"  {code}: File not found: {filepath}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Parse all Index compteur sheets
    daily_index = []
    # Parse all Prod solaire sheets
    daily_prod = []

    for sn in wb.sheetnames:
        sn_lower = sn.lower()
        month = match_month(sn)
        if month is None:
            continue

        if "index" in sn_lower:
            entries = parse_index_sheet(wb[sn])
            daily_index.extend(entries)
        elif "prod" in sn_lower:
            entries = parse_prod_solaire_sheet(wb[sn])
            daily_prod.extend(entries)

    # Merge data: prefer Prod solaire for availability/interruptions, Index for irradiance/peak
    index_by_date = {r["date"]: r for r in daily_index}
    prod_by_date = {r["date"]: r for r in daily_prod}
    all_dates = sorted(set(list(index_by_date.keys()) + list(prod_by_date.keys())))

    daily = []
    for d in all_dates:
        idx = index_by_date.get(d, {})
        prd = prod_by_date.get(d, {})

        entry = {"date": d}
        # Production: prefer prod solaire (from meter), fallback to index
        entry["prodKwh"] = prd.get("prodKwh") or idx.get("prodKwh") or 0
        entry["deliveredKwh"] = prd.get("deliveredKwh") or entry["prodKwh"]
        entry["consumedKwh"] = prd.get("consumedKwh", 0)
        entry["peakKw"] = idx.get("peakKw", 0)
        entry["irradiance"] = idx.get("irradiance", 0)
        entry["availHours"] = prd.get("availHours", 0)
        entry["unschedInterrupt"] = prd.get("unschedInterrupt", 0)
        entry["schedInterrupt"] = prd.get("schedInterrupt", 0)
        entry["couplingHours"] = idx.get("couplingHours", 0)

        if entry["prodKwh"] > 0:
            daily.append(entry)

    # Aggregate by month
    monthly = {}
    for d in daily:
        ym = d["date"][:7]  # "2026-01"
        if ym not in monthly:
            monthly[ym] = {
                "month": ym,
                "totalProdKwh": 0,
                "totalDeliveredKwh": 0,
                "totalConsumedKwh": 0,
                "maxPeakKw": 0,
                "avgIrradiance": 0,
                "totalAvailHours": 0,
                "totalUnschedInterrupt": 0,
                "daysWithData": 0,
                "dailyProd": [],
            }
        m = monthly[ym]
        m["totalProdKwh"] += d["prodKwh"]
        m["totalDeliveredKwh"] += d["deliveredKwh"]
        m["totalConsumedKwh"] += d["consumedKwh"]
        if d["peakKw"] > m["maxPeakKw"]:
            m["maxPeakKw"] = d["peakKw"]
        m["avgIrradiance"] += d["irradiance"]
        m["totalAvailHours"] += d["availHours"]
        m["totalUnschedInterrupt"] += d["unschedInterrupt"]
        m["daysWithData"] += 1
        m["dailyProd"].append(round(d["prodKwh"], 1))

    # Finalize monthly
    for ym, m in monthly.items():
        if m["daysWithData"] > 0:
            m["avgIrradiance"] = round(m["avgIrradiance"] / m["daysWithData"], 2)
            m["avgDailyProdKwh"] = round(m["totalProdKwh"] / m["daysWithData"], 1)
        else:
            m["avgDailyProdKwh"] = 0
        m["totalProdKwh"] = round(m["totalProdKwh"], 1)
        m["totalDeliveredKwh"] = round(m["totalDeliveredKwh"], 1)
        m["totalConsumedKwh"] = round(m["totalConsumedKwh"], 1)
        m["maxPeakKw"] = round(m["maxPeakKw"])
        m["totalAvailHours"] = round(m["totalAvailHours"], 1)
        m["totalUnschedInterrupt"] = round(m["totalUnschedInterrupt"], 1)

    # Latest day
    latest = daily[-1] if daily else None
    latest_date = latest["date"] if latest else None

    # Overall stats
    total_prod_kwh = sum(d["prodKwh"] for d in daily)
    total_days = len(daily)
    avg_daily = round(total_prod_kwh / total_days, 1) if total_days > 0 else 0
    max_peak = max((d["peakKw"] for d in daily), default=0)

    site_data = {
        "code": code,
        "name": site_cfg["name"],
        "entity": site_cfg["entity"],
        "centrale": site_cfg["centrale"],
        "loc": site_cfg["loc"],
        "capacityKwc": site_cfg["capacityKwc"],
        "capacityMw": round(site_cfg["capacityKwc"] / 1000, 1),
        "latestDate": latest_date,
        "totalProdKwh": round(total_prod_kwh, 1),
        "totalDays": total_days,
        "avgDailyKwh": avg_daily,
        "maxPeakKw": round(max_peak),
        "monthly": sorted(monthly.values(), key=lambda x: x["month"]),
        "daily": daily,
    }

    return site_data


def build_enr_sites():
    """Build and return ENR sites data dict (for API use). Returns {"ENR_SITES": [...]}."""
    all_sites = []
    for code, cfg in SITES.items():
        data = build_site_data(code, cfg)
        if data:
            all_sites.append(data)
    return {"ENR_SITES": all_sites}


def generate():
    print("Generating EnR site data...")
    print(f"Source: {ENR_DIR}")

    all_sites = []
    for code, cfg in SITES.items():
        data = build_site_data(code, cfg)
        if data:
            all_sites.append(data)
            m_summary = ", ".join(
                f"{m['month']}={m['totalProdKwh']/1000:.1f}MWh({m['daysWithData']}d)"
                for m in data["monthly"]
            )
            print(
                f"  {code}: {data['capacityKwc']}kWc, "
                f"{data['totalDays']} days, "
                f"avg={data['avgDailyKwh']/1000:.1f}MWh/d, "
                f"peak={data['maxPeakKw']}kW"
            )
            print(f"    Monthly: {m_summary}")
        else:
            print(f"  {code}: No data")

    # Output JS
    js = "// Auto-generated from EnR production Excel files\n"
    js += f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    js += f"const ENR_SITES = {json.dumps(all_sites, ensure_ascii=False)};\n"

    outpath = os.path.join(os.path.dirname(__file__), "enr_site_data.js")
    with open(outpath, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"\nGenerated enr_site_data.js ({len(js)} bytes, {len(all_sites)} sites)")


if __name__ == "__main__":
    generate()
